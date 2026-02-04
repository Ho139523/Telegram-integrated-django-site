#functions.py
from email import message
import profile
from pydoc import describe

from click import command
from markupsafe import Markup
# from tkinter.font import names
from utils.variables.TOKEN import TOKEN, BOT_ID
import requests
import subprocess
import re
from telebot import TeleBot
from telebot.types import Message
from telebot.storage import StateMemoryStorage
from accounts.models import ProfileModel, Address
from products.models import Product, Category, ProductImage, ProductAttribute, Store, ProductVariant
from payment.models import Transaction, Sale, Cart, CartItem
import functools
from telbot.models import CachedMedia
import os
from django.conf import settings
from AI.settings import current_site as settings_current_site
import requests
from django.core.files.base import ContentFile
import json
import urllib.parse
import base64
import uuid
from django.core.cache import cache
from utils.variables.translate import translations
from django.db.models.functions import Lower
import itertools
import traceback
import re
from telebot.types import Message
from products.models import (
    Product, ProductAttribute, ProductImage,
    ProductVariant, ProductOption, ProductOptionValue
)
from accounts.models import ProfileModel
from products.models import Store, Category
from django.db.models import Q
from django.core.exceptions import ValidationError
from django.conf import settings
from pathlib import Path
from asgiref.sync import sync_to_async
from collections import defaultdict

# send_product_message function
from telebot import types


# bot settings
from telebot.storage import StateMemoryStorage
state_storage = StateMemoryStorage()
app = TeleBot(token=TOKEN, state_storage=state_storage)


from telbot.sessions import CartSessionManager, RedisStateManager, SessionManager, RedisExportManager

# Access shared user_sessions
from telbot.sessions import session_manager

from utils.telbot.variables import *
import os
import requests
from django.conf import settings

from utils.telbot.variables import home_menu

from PIL import Image, ImageDraw, ImageFont
from django.utils import timezone

from django.conf import settings as sett


from utils.funcs.geonames_address import get_country_choices, get_province_choices, get_city_choices
from pathlib import Path
from AI.settings import MEDIA_URL


def t(msg, key, chat_id=None, **kwargs):
    try:
        if isinstance(msg, types.Message):
            message = msg
        elif isinstance(msg, types.CallbackQuery):
            message = msg.message
        else:
            message = None

        if chat_id:
            print(chat_id)
            pass
        else:
            chat_id = message.chat.id

        lang = ProfileModel.objects.get(tel_id=chat_id).lang
        text = translations.get(key, {}).get(lang, translations[key]["en"])
        
        # جایگذاری متغیرها
        if kwargs:
            text = text.format(**kwargs)
        return text
    except Exception as e:
        print(traceback.format_exc())




def get_tunnel_password():
    try:
        result = subprocess.run(
            ["curl", "-s", "https://loca.lt/mytunnelpassword"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        if result.returncode == 0:
            password = result.stdout.strip()  # حذف فاصله‌ها و خط‌های اضافی
            return password
        else:
            print("Error fetching password:", result.stderr)
            return None
    except Exception as e:
        print(f"An error occurred: {traceback.format_exc()}")
        return None



# Getting website address and webhook

def get_current_webhook(TOKEN=TOKEN):
    bot_token = TOKEN  # Ensure you have your bot token in Django settings
    response = requests.get(f'https://api.telegram.org/bot{bot_token}/getWebhookInfo')

    if response.status_code == 200:
        webhook_info = response.json()

        # Check if there's a URL set for the webhook
        if webhook_info.get('ok') and webhook_info['result'].get('url'):
            return webhook_info['result']['url']
        else:
            return "No webhook URL set."
    else:
        return "Failed to retrieve webhook info."

def get_current_site(TOKEN=TOKEN):
    bot_token = TOKEN  # Ensure you have your bot token in Django settings
    response = requests.get(f'https://api.telegram.org/bot{bot_token}/getWebhookInfo')

    if response.status_code == 200:
        site_info = response.json()

        # Check if there's a URL set for the webhook
        if site_info.get('ok') and site_info['result'].get('url'):
            return site_info['result']['url'][:-9]
        else:
            return "No site URL set."
    else:
        return "Failed to retrieve site info."


# بررسی معتبر بودن رمز عبور
def validate_password(password):
    # شرط طول رمز عبور حداقل ۸ کاراکتر
    if len(password) < 8:
        return False, "رمز عبور باید حداقل ۸ کاراکتر باشد."

    # شرط حروف کوچک
    if not re.search(r"[a-z]", password):
        return False, "رمز عبور باید حداقل شامل یک حرف کوچک باشد."

    # شرط حروف بزرگ
    if not re.search(r"[A-Z]", password):
        return False, "رمز عبور باید حداقل شامل یک حرف بزرگ باشد."

    # شرط عدد
    if not re.search(r"[0-9]", password):
        return False, "رمز عبور باید حداقل شامل یک عدد باشد."

    # شرط علامت‌ها
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
        return False, "رمز عبور باید حداقل شامل یک علامت باشد."

    # اگر همه شرایط برقرار بود
    return True, "رمز عبورت خوبه."


def validate_username(username):
    # Check length
    if len(username) < 5 or len(username) > 32:
        return False, "طول نام کاربری باید بین 5 تا 32 حرف باشد."

    # Check for allowed characters and disallow "."
    if not re.match(r"^[a-zA-Z0-9_]+$", username):
        return False, "نام کاربری تنها شامل حروف، عدد و underline باشد."

    # Check for presence of "."
    if "." in username:
        return False, "نام کاربری نمی تواند شامل «.» باشد."

    return True, "این نام کاربری خوبه"



############################  SEND PRODUCT MESSAGE  ############################

class UserOrderManager:
    """مدیریت تعداد سفارش‌های هر کاربر"""
    def __init__(self):
        self.user_counts = {}

    def increase(self, chat_id):
        self.user_counts[chat_id] = self.user_counts.get(chat_id, 0) + 1

    def decrease(self, chat_id):
        if self.user_counts.get(chat_id, 0) > 1:
            self.user_counts[chat_id] -= 1
        else:
            self.user_counts.pop(chat_id, None)

    def get_count(self, chat_id):
        return self.user_counts.get(chat_id, 0)





# نحوه استفاده:
# product_handler = ProductHandler(app, product, current_site)
# product_handler.send_product_message(chat_id)

############################  PAGINATION  ############################

import math
import json
import redis

class InlineKeyboardPaginator:
    def __init__(
            self,
            user_id,
            items=None,
            per_page=10,
            row_size=3,
            remember_last_page=False,
            redis_host="localhost",
            redis_port=6379
    ):
        """
        :param user_id: آیدی کاربر
        :param items: لیست آیتم‌ها (مثل کشورها) – اگر None باشد از Redis بارگذاری می‌شود
        :param per_page: تعداد آیتم در هر صفحه
        :param row_size: تعداد دکمه در هر ردیف
        :param remember_last_page: اگر True باشد، صفحه‌ی آخر کاربر ذخیره می‌شود
        """
        self.user_id = user_id
        self.per_page = per_page
        self.row_size = row_size
        self.remember_last_page = remember_last_page
        self.redis = redis.Redis(host=redis_host, port=redis_port, db=0, decode_responses=True)

        if items is not None:
            self.items = items
            self.total_pages = max(1, math.ceil(len(items) / per_page))
            self._save_state(1 if not remember_last_page else self._get_saved_page(), items)
        else:
            self._load_state()

    def _get_key(self):
        return f"paginator:{self.user_id}"

    def _save_state(self, page, items):
        """ذخیره وضعیت در Redis"""
        data = {
            "items": items,
            "page": page,
            "per_page": self.per_page,
            "row_size": self.row_size,
            "remember_last_page": self.remember_last_page
        }
        self.redis.set(self._get_key(), json.dumps(data))

    def _load_state(self):
        """لود وضعیت از Redis"""
        data = self.redis.get(self._get_key())
        if not data:
            raise ValueError(f"No paginator state found for user_id={self.user_id}")
        data = json.loads(data)
        self.items = data["items"]
        self.per_page = data.get("per_page", self.per_page)
        self.row_size = data.get("row_size", self.row_size)
        self.remember_last_page = data.get("remember_last_page", self.remember_last_page)
        self.total_pages = max(1, math.ceil(len(self.items) / self.per_page))
        self.set_page(data.get("page", 1))

    def _get_saved_page(self):
        """بررسی صفحه ذخیره شده (اگر وجود داشته باشد)"""
        data = self.redis.get(self._get_key())
        if data:
            data = json.loads(data)
            return data.get("page", 1)
        return 1

    def get_current_page(self):
        data = json.loads(self.redis.get(self._get_key()))
        return data.get("page", 1)

    def set_page(self, page: int):
        page = max(1, min(self.total_pages, page))
        data = json.loads(self.redis.get(self._get_key()))
        if self.remember_last_page:  # Only update saved page if this option is enabled
            data["page"] = page
        else:
            data["page"] = 1  # Always reset to 1 when user leaves
        self.redis.set(self._get_key(), json.dumps(data))

    def next_page(self):
        current = self.get_current_page()
        if current < self.total_pages:
            self.set_page(current + 1)

    def prev_page(self):
        current = self.get_current_page()
        if current > 1:
            self.set_page(current - 1)

    def get_buttons_for_sendmarkup(self):
        try:
            current_page = self.get_current_page()
            start = (current_page - 1) * self.per_page
            end = start + self.per_page
            current_items = self.items[start:end]

            buttons = {}
            button_layout = []

            row_count = 0
            for idx, item in enumerate(current_items, start=1):
                buttons[item] = {"callback_data": item, "index": idx}
                row_count += 1
                if row_count == self.row_size or idx == len(current_items):
                    button_layout.append(row_count)
                    row_count = 0


            if len(self.items) <= self.per_page:
                return buttons, button_layout

            control_buttons = []
            idx = len(buttons) + 1
            if current_page > 1:
                buttons["⬅️ قبلی"] = {"callback_data": "prev", "index": idx}
                control_buttons.append("⬅️ قبلی")
                idx += 1

            buttons[f"{current_page}/{self.total_pages}"] = {"callback_data": "current", "index": idx}
            control_buttons.append(f"{current_page}/{self.total_pages}")
            idx += 1

            if current_page < self.total_pages:
                buttons["بعدی ➡️"] = {"callback_data": "next", "index": idx}
                control_buttons.append("بعدی ➡️")

            button_layout.append(len(control_buttons))
            return buttons, button_layout
        except Exception as e:
            print(f"errors in class: {traceback.format_exc()}")

    @staticmethod
    def load_from_redis(user_id, redis_host="localhost", redis_port=6379):
        return InlineKeyboardPaginator(
            user_id=user_id,
            items=None,
            redis_host=redis_host,
            redis_port=redis_port
        )

############################  MEASURE PERFORMANCE  ############################

import time
import functools
from django.db import connection
from django.db import reset_queries

def measure_performance(func):
    """
    دکوراتور برای اندازه‌گیری سرعت اجرای توابع و تعداد کوئری‌های دیتابیس
    """
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        # زمان شروع
        start_time = time.time()
        
        # ریست کردن شمارشگر کوئری‌ها
        reset_queries()
        
        try:
            # اجرای تابع اصلی
            result = func(*args, **kwargs)
            
            # زمان پایان
            end_time = time.time()
            execution_time = end_time - start_time
            
            # تعداد کوئری‌های اجرا شده
            num_queries = len(connection.queries)
            
            # اطلاعات عملکرد
            performance_info = {
                'function_name': func.__name__,
                'execution_time': round(execution_time, 4),
                'num_queries': num_queries,
                'queries': connection.queries if num_queries > 0 else []
            }
            
            # چاپ اطلاعات
            print(f"🚀 PERFORMANCE REPORT - {func.__name__}")
            print(f"⏱️  Execution time: {execution_time:.4f} seconds")
            print(f"📊 Number of database queries: {num_queries}")
            
            if num_queries > 0:
                print("🔍 Queries executed:")
                for i, query in enumerate(connection.queries, 1):
                    print(f"   {i}. {query['sql']} (Time: {query['time']}s)")
            
            print("=" * 60)
            
            return result
            
        except Exception as e:
            # در صورت خطا هم زمان را اندازه بگیریم
            end_time = time.time()
            execution_time = end_time - start_time
            
            print(f"❌ ERROR in {func.__name__}")
            print(f"⏱️  Execution time before error: {execution_time:.4f} seconds")
            print(f"💥 Error: {traceback.format_exc()}")
            print("=" * 60)
            
            raise e
    
    return wrapper

def add_performance_monitoring_to_class(cls):
    """
    اضافه کردن مانیتورینگ عملکرد به همه متدهای یک کلاس

    استفاده:
    @add_performance_monitoring_to_class
    class ProductHandler:
    تمام متدهای شما به صورت خودکار مانیتور می‌شوند
    pass
    """
    for attr_name in dir(cls):
        attr = getattr(cls, attr_name)
        if callable(attr) and not attr_name.startswith('_'):
            setattr(cls, attr_name, measure_performance(attr))
    return cls


############################  SEND MARKUP  ############################

from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from telebot import types

import threading
from django.core.cache import cache

class SendMarkup:
    def __init__(self, bot, chat_id, text=None, buttons=None, button_layout=None, handlers=None, message=None):
        self.bot = bot
        self.chat_id = chat_id
        self.text = text
        self.buttons = buttons or []
        self.button_layout = button_layout or []
        self.handlers = handlers or {}
        self._keyboard_cache = None
        self.message = message

    def _validate_button(self, text, callback_data, is_url=False):
        """اعتبارسنجی دکمه قبل از ساخت"""
        if not text or not isinstance(text, str) or text.strip() == "":
            return False, "متن دکمه نمی‌تواند خالی باشد"
        
        # برای دکمه‌های URL، callback_data می‌تواند خالی باشد
        if not is_url and (not callback_data or callback_data.strip() == ""):
            return False, "callback_data نمی‌تواند خالی باشد"
            
        if not isinstance(callback_data, str):
            return False, "callback_data باید رشته باشد"
            
        return True, "معتبر"


    def _convert_buttons_to_list(self):
        """تبدیل دکمه‌ها به فرمت لیست یکپارچه با اعتبارسنجی"""
        if not self.buttons:
            return []
            
        button_list = []
        
        try:
            # اگر buttons از قبل لیست است
            if isinstance(self.buttons, list):
                for item in self.buttons:
                    if len(item) >= 3:  # (text, callback_data, index)
                        text, callback_data, index = item[0], item[1], item[2]
                        is_valid, message = self._validate_button(text, callback_data)
                        if is_valid:
                            button_list.append((text, callback_data, index))
                        else:
                            print(f"دکمه نامعتبر حذف شد: {text} - {message}")
                    else:
                        print(f"فرمت دکمه نامعتبر: {item}")
            
            # اگر buttons دیکشنری است
            elif isinstance(self.buttons, dict):
                for text, button_data in self.buttons.items():
                    callback_data = ""
                    url = ""
                    index = len(button_list) + 1
                    
                    if isinstance(button_data, dict):
                        # فرمت جدید با دیکشنری
                        callback_data = button_data.get('callback_data', '')
                        url = button_data.get('url', '')  # اضافه کردن پشتیبانی از URL
                        index = button_data.get('index', index)
                        
                        # اولویت با URL است اگر وجود دارد
                        if url:
                            callback_data = url  # استفاده از URL به عنوان callback_data
                    elif isinstance(button_data, (list, tuple)) and len(button_data) >= 2:
                        # فرمت قدیمی با تاپل
                        callback_data, index = button_data[0], button_data[1]
                    else:
                        print(f"فرمت دکمه نامعتبر برای {text}: {button_data}")
                        continue
                    
                    # اعتبارسنجی: اگر URL داریم، callback_data خالی مجاز است
                    if url:
                        # برای دکمه‌های URL، callback_data می‌تواند خالی باشد
                        if not text or not isinstance(text, str) or text.strip() == "":
                            print(f"دکمه نامعتبر حذف شد: {text} - متن دکمه نمی‌تواند خالی باشد")
                            continue
                        button_list.append((text, url, index))
                    else:
                        # برای دکمه‌های معمولی، callback_data باید پر باشد
                        is_valid, message = self._validate_button(text, callback_data)
                        if is_valid:
                            button_list.append((text, callback_data, index))
                        else:
                            print(f"دکمه نامعتبر حذف شد: {text} - {message}")
            
            else:
                print(f"فرمت buttons نامعتبر: {type(self.buttons)}")
                
        except Exception as e:
            print(f"خطا در تبدیل دکمه‌ها: {traceback.format_exc()}")
            
        return button_list

        
    def generate_keyboard(self):
        """ساخت کیبورد با اعتبارسنجی کامل"""
        if self._keyboard_cache:
            return self._keyboard_cache
            
        markup = types.InlineKeyboardMarkup()
        
        if not self.buttons:
            return markup

        # تبدیل به فرمت لیست یکپارچه
        button_list = self._convert_buttons_to_list()
        
        if not button_list:
            print("هیچ دکمه معتبری برای نمایش وجود ندارد")
            return markup
        
        # مرتب‌سازی بر اساس ایندکس
        try:
            sorted_buttons = sorted(button_list, key=lambda x: x[2])
        except Exception as e:
            print(f"خطا در مرتب‌سازی دکمه‌ها: {traceback.format_exc()}")
            sorted_buttons = button_list

        # ساخت دکمه‌های اینلاین
        inline_buttons = []
        for text, callback_data, index in sorted_buttons:
            try:
                # بررسی اگر دکمه لینک باشد (شامل http یا https)
                if callback_data.startswith(('http://', 'https://')):
                    inline_buttons.append(types.InlineKeyboardButton(text, url=callback_data))
                else:
                    inline_buttons.append(types.InlineKeyboardButton(text, callback_data=callback_data))
            except Exception as e:
                print(f"خطا در ساخت دکمه {text}: {traceback.format_exc()}")
                continue

        if not inline_buttons:
            print("هیچ دکمه اینلاین معتبری ساخته نشد")
            return markup

        # چیدمان دکمه‌ها بر اساس طرح‌بندی
        try:
            index = 0
            for row_size in self.button_layout:
                if index >= len(inline_buttons):
                    break
                    
                if row_size <= 0:
                    print(f"سایز ردیف نامعتبر: {row_size}")
                    continue
                    
                row_buttons = inline_buttons[index:index + row_size]
                if row_buttons:  # اطمینان از خالی نبودن ردیف
                    markup.row(*row_buttons)
                index += row_size
                
            # اگر دکمه‌های باقیمانده داریم، آن‌ها را در ردیف آخر قرار دهیم
            if index < len(inline_buttons):
                remaining_buttons = inline_buttons[index:]
                markup.row(*remaining_buttons)
                
        except Exception as e:
            print(f"خطا در چیدمان دکمه‌ها: {traceback.format_exc()}")

        self._keyboard_cache = markup
        return markup

    def send(self):
        """ارسال پیام با هندل خطا"""
        try:
            markup = self.generate_keyboard()
            self.bot.send_message(
                chat_id=self.chat_id,
                text=self.text,
                reply_markup=markup,
                parse_mode="HTML"
            )
        except Exception as e:
            print(f"Error in SendMarkup.send: {traceback.format_exc()}")
            # تلاش برای ارسال بدون دکمه در صورت خطا
            try:
                self.bot.send_message(
                    chat_id=self.chat_id,
                    text=self.text,
                    parse_mode="HTML"
                )
            except Exception as e2:
                print(f"Error sending without buttons: {e2}")

    def edit(self, message_id):
        """ویرایش هوشمند پیام (text یا caption)"""
        try:
            markup = self.generate_keyboard()

            # اگر message داریم و پیام photo است → caption
            if self.message and getattr(self.message, "content_type", None) == "photo":
                self.bot.edit_message_caption(
                    chat_id=self.chat_id,
                    message_id=message_id,
                    caption=self.text,
                    reply_markup=markup,
                    parse_mode="HTML"
                )
            else:
                # حالت پیش‌فرض: text
                self.bot.edit_message_text(
                    chat_id=self.chat_id,
                    message_id=message_id,
                    text=self.text,
                    reply_markup=markup,
                    parse_mode="HTML"
                )

        except Exception as e:
            if "message is not modified" not in str(e):
                print(f"Error in SendMarkup.edit: {traceback.format_exc()}")


    def handle_callback(self, call):
        """مدیریت کلیک روی دکمه‌ها"""
        callback_data = call.data
        if callback_data in self.handlers:
            try:
                self.handlers[callback_data](call)
            except Exception as e:
                print(f"Error in handler for {callback_data}: {traceback.format_exc()}")

    def debug_buttons(buttons):
        """تابع کمکی برای دیباگ دکمه‌ها"""
        print("=== DEBUG BUTTONS ===")
    
        if not buttons:
            print("دکمه‌ها خالی هستند")
            return
        
        if isinstance(buttons, list):
            print(f"فرمت: لیست ({len(buttons)} آیتم)")
            for i, item in enumerate(buttons):
                print(f"  {i}: {item}")
        elif isinstance(buttons, dict):
            print(f"فرمت: دیکشنری ({len(buttons)} آیتم)")
            for key, value in buttons.items():
                print(f"  '{key}': {value}")
        else:
            print(f"فرمت نامشخص: {type(buttons)}")
    
        print("=====================")


############################  CHECK SUBSCRIPTION  ############################

import logging
from utils.variables.CHANNELS import my_channels_with_atsign, my_channels_without_atsign

logger = logging.getLogger(__name__)

class SubscriptionClass:
    def __init__(self, bot: TeleBot):
        self.bot = bot
        self.my_channels_with_atsign = my_channels_with_atsign
        self.my_channels_without_atsign = my_channels_without_atsign
        self.current_site = settings_current_site

    def handle_check_subscription(self, call: types.CallbackQuery):
        """✅ بررسی عضویت هنگام کلیک روی دکمه 'عضو شدم'"""
        chat_id = call.message.chat.id
        user_id = call.from_user.id



        # بررسی عضویت
        is_member = self.check_subscription(user_id)

        if is_member:

            try:

                # ✅ پاسخ اولیه به Callback Query
                self.bot.answer_callback_query(call.id, t(call.message, "checking_membership"), show_alert=False)
                self.bot.edit_message_text(t(call.message, "membership_confirmed"),
                                           chat_id=chat_id, message_id=call.message.message_id)

                from telbot.views import start
                start(call.message)
            except Exception as e:
                self.bot.send_message(user_id, f"error is: {traceback.format_exc()}")
        else:
            self.bot.answer_callback_query(call.id, t(call.message, "not_subscribed"), show_alert=True)

    def register_handlers(self):
        """🔹 ثبت هندلرهای مورد نیاز"""
        self.bot.callback_query_handler(func=lambda call: call.data == "check_subscription2")(self.handle_check_subscription)

    def check_subscription(self, user, channels=None):
        """✅ بررسی می‌کند که کاربر در کانال عضو شده است یا نه"""
        if channels is None:
            channels = self.my_channels_with_atsign
        for channel in channels:
            try:
                is_member = self.bot.get_chat_member(chat_id=channel, user_id=user)
                if is_member.status in ["kicked", "left"]:
                    return False
            except Exception as e:
                
                logger.error(f"🚨 خطا در بررسی عضویت کاربر {user} در کانال {channel}: {traceback.format_exc()}")
                return False
        return True

    def subscription_offer(self, message):
        """❌ اگر کاربر عضو نباشد، دکمه‌های عضویت نمایش داده شوند"""
        channel_markup = types.InlineKeyboardMarkup()
        check_subscription_button = types.InlineKeyboardButton(text=t(message, "subscribed"), callback_data='check_subscription2')
        channel_subscription_button = types.InlineKeyboardButton(text=t(message, "join_channel"), url=f"https://t.me/{self.my_channels_without_atsign[0]}")
        group_subscription_button = types.InlineKeyboardButton(text=t(message, 'join_group'), url=f"https://t.me/{self.my_channels_without_atsign[1]}")

        channel_markup.add(channel_subscription_button, group_subscription_button)
        channel_markup.add(check_subscription_button)

        if not self.check_subscription(user=message.chat.id):
            self.bot.send_message(message.chat.id, t(message, "verify_membership"), reply_markup=channel_markup)
            return False
        return True

subscription = SubscriptionClass(app)

############################  SEND MENU  ############################

# Helper function to send menu
def send_menu(message, options, current_menu, extra_buttons=None, cols=3, extra_cols=2):
    """Send a translated menu with options and update the session."""

    if subscription.subscription_offer(message):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

        # زبان کاربر
        profile = ProfileModel.objects.get(tel_id=message.chat.id)
        lang = profile.lang if profile.lang else "fa"

        # ترجمه گزینه‌های اصلی
        rows = [options[i:i + cols] for i in range(0, len(options), cols)]
        for row in rows:
            translated_row = [
                translations.get(key, {}).get(lang, translations.get(key, {}).get("en", key))
                for key in row
            ]
            markup.row(*translated_row)

        # ترجمه گزینه‌های اضافه
        if extra_buttons:
            extra_rows = [extra_buttons[i:i + extra_cols] for i in range(0, len(extra_buttons), extra_cols)]
            for extra_row in extra_rows:
                translated_row = [
                    translations.get(key, {}).get(lang, translations.get(key, {}).get("en", key))
                    for key in extra_row
                ]
                markup.row(*translated_row)

        return markup


############################  SEE PRODUCTS  ############################

# Top discounts
def handle_products(message):
    if subscription.subscription_offer(message):
        chat_id = message.chat.id
        subcategory = message.text
        options = [t(message, "most_selling"), t(message, "most_expensive"), t(message, "cheapest"), t(message, "most_discounted")]

        markup = send_menu(message, options, "products", retun_menue)
        session = session_manager.get_user_session(chat_id, namespace="menu")
        current_category = Category.objects.get(title__iexact=session["current_menu"], status=True)
        app.send_message(chat_id, f"{current_category.get_full_path()}", reply_markup=markup)


############################  CATEGORY MENU  ############################

class CategoryClass:

    def __init__(self):
        pass

    def handle_category(self, message):
        if subscription.subscription_offer(message):
            try:
                extra_menu = [t(message, "cancel_action")]
                session = session_manager.get_user_session(message.chat.id, namespace="menu")
                print(f"begining of handle_category: {session}")
                profile = ProfileModel.objects.get(tel_id=message.chat.id)
                if profile.seller_mode:
                    store = Store.objects.get(owner=profile)
                else:
                    store = profile.server_store
                if not session.get("category") and not session.get("product") and not store.categories.exists():
                    # Store has no categories
                    app.send_message(message.chat.id, t(message, "store_empty"))
                    return
                    
                    # do something...
                if session.get("category") and session.get("menu_delete") and session.get("delete_sure"):
                    text = t(message, "category_deleted_successfully")
                    session["delete_sure"] = False
                    session_manager.set_user_session(message.chat.id, session, namespace="menu")
                elif session.get("category") and session.get("menu_delete"):
                    if not store.categories.exists():
                        app.send_message(message.chat.id, t(message, "no_category_to_delete"))
                        session["menu_delete"] = False
                        session_manager.set_user_session(message.chat.id, session, namespace="menu")
                        return
                    text = t(message, "delete_category_title_prompt") + "\n\n" + t(message, "delete_category_warning")
                elif session.get("category") and session.get("menu_add"):
                    if not store.categories.exists():
                        self.add_category(message)
                        return
                    text = t(message, "add_subcategory_select_parent")
                elif session.get("category") and session.get("category_deactivate"):
                    text = t(message, "choose_category_toggle")
                elif session.get("product"):
                    text = t(message, "select_subcategory_for_product")
                elif not session.get("category") and not session.get("product"):
                    extra_menu = home_menu
                    button = [t(message, "delete_category_and_subcategories"), ]
                    for b in button:
                        extra_menu.remove(b) if b in extra_menu else None
                    text = t(message, "product_category_question")
                else:
                    text = t(message, "category_deleted_successfully")
                cats = Category.objects.filter(parent__isnull=True, status=True, store=store).values_list('title', flat=True)
                if session.get("category") and session.get("category_deactivate"):
                    cats = Category.objects.filter(parent__isnull=True, store=store).values_list('title', flat=True)
                markup = send_menu(message, cats, message.text, extra_menu, extra_cols=1)
                app.send_message(message.chat.id, text, reply_markup=markup)
            except Exception as e:
                app.send_message(message.chat.id, "خطایی رخ داده است. لطفاً دوباره تلاش کنید.")
                print(f"Error: {traceback.format_exc()}")

    def handle_subcategory(self, message):
        try:
            if subscription.subscription_offer(message):
                session = session_manager.get_user_session(message.chat.id, namespace="menu")
                session["current_menu"] = message.text.lower()
                session_manager.set_user_session(message.chat.id, session, namespace="menu")

                extra_menu = ["🔙", t(message, "cancel_action")]
                
                if session.get("category") and session.get("category_deactivate"):
                    current_category = Category.objects.get(title__iexact=message.text.title())
                    children = [child.title for child in current_category.get_next_layer_categories(both=True)]

                # when you want to delete you must be able to delete even deactivated cats
                elif session.get("category") and session.get("menu_delete"):
                    current_category = Category.objects.get(title__iexact=message.text.title())
                    children = [child.title for child in current_category.get_next_layer_categories(both=True)]
                
                else:
                    current_category = Category.objects.get(title__iexact=message.text.title(), status=True)
                    children = [child.title for child in current_category.get_next_layer_categories()]

                
                print(f"begining of handle_subcategory: {session}")
                

                if children == []:
                    if session.get("category") and session.get("menu_delete"):
                        self.delete_sure(message)
                        try:
                            message.text = current_category.get_parents()[0].title
                        except:
                            pass
                    elif session.get("category") and session.get("menu_add"):
                        self.add_category(message)
                        try:
                            print(current_category)
                            print(current_category.get_parents())
                            message.text = current_category.get_parents()[0].title
                        except:
                            pass
                    elif session.get("category") and session.get("category_deactivate"):
                        self.deactivate_category_sure(message)
                        # message.text = current_category.get_parents()[0].title
                    else:
                        if session.get('product_cat_selection'):
                            a = ProductBot(app)
                            a.get_category(message)
                        else:
                            fake_message = message
                            fake_message.text = "hi"
                            handle_products(fake_message)
                else:
                    if not session.get("category") and not session.get("product"):
                        extra_menu = retun_menue
                    elif session.get("category"):
                        if session.get("menu_delete"):
                            button = t(message, "delete_category_and_subcategories")
                            extra_menu.append(button) if button not in extra_menu else extra_menu
                        elif session.get("category_deactivate"):
                            if not [child.title for child in current_category.get_next_layer_categories(status=True)]:
                                button = t(message, "activate_category")
                            else:
                                button = t(message, "deactivate_category")
                            extra_menu.append(button) if button not in extra_menu else extra_menu
                    if session.get("category") and session.get("menu_delete") and session.get("delete_sure"):
                        text = t(message, "category_deleted_successfully")
                        session["delete_sure"] = False
                        session_manager.set_user_session(message.chat.id, session, namespace="menu")
                    elif session.get("category") and session.get("menu_delete"):
                        text = f"{Category.objects.get(title__iexact=session['current_menu'], status=True).get_full_path()}"
                    elif session.get("category") and session.get("category_deactivate"):
                        text = f"{Category.objects.get(title__iexact=session['current_menu']).get_full_path()}"
                    else:
                        button = [t(message, "delete_category_and_subcategories"), t(message, "deactivate_category")]
                        for b in button:
                            extra_menu.remove(b) if b in extra_menu else None
                        text = f"{Category.objects.get(title__iexact=session['current_menu'], status=True).get_full_path()}"
                    markup = send_menu(message, children, message.text, extra_menu)
                    app.send_message(message.chat.id, text, reply_markup=markup)
        except Exception as e:
            print(f"Error: {traceback.format_exc()}")

    def add_category(self, message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="menu")
            extra_menu = [t(message, "cancel_action")]

            # Mark we are waiting for new category title
            session["get_new_category"] = True

            # IMPORTANT: Only set current_menu if we are actually inside a category
            if session.get("current_menu"):
                # user is inside a category, so add to that category
                session["parent_for_new"] = session["current_menu"]
            else:
                # user is in root menu
                session["parent_for_new"] = None  

            button = [t(message, "delete_category_and_subcategories"), t(message, "deactivate_category")]
            for b in button:
                if b in extra_menu:
                    extra_menu.remove(b)

            markup = send_menu(message, [], 'cat_delete_sure', extra_menu)
            session_manager.set_user_session(message.chat.id, session, namespace="menu")

            app.send_message(message.chat.id, t(message, "enter_new_category_title"), reply_markup=markup)

        except Exception as e:
            print(traceback.format_exc())


    def deactivate_category_sure(self, message):
        try:
            extra_menu = [t(message, "cancel_action")]
            session = session_manager.get_user_session(message.chat.id, namespace="menu")
            session["deactivate_category_sure"] = True
            markup = send_menu(message, [t(message, "yes_im_sure"), t(message, "cancel_action")], 'cat_delete_sure')
            cat = Category.objects.get(title__iexact=session.get("current_menu"), store__owner__tel_id=message.chat.id)
            if cat.status:
                app.send_message(message.chat.id, t(message, "confirm_deactivate_category"), reply_markup=markup)
            else:
                app.send_message(message.chat.id, t(message, "confirm_activate_category"), reply_markup=markup)
            session_manager.set_user_session(message.chat.id, session, namespace="menu")
        except Exception as e:
            print(traceback.format_exc())

    def delete_sure(self, message):
        try:
            extra_menu = [t(message, "cancel_action")]
            session = session_manager.get_user_session(message.chat.id, namespace="menu")
            session["delete_sure"] = True
            markup = send_menu(message, [t(message, "yes_im_sure"), t(message, "cancel_action")], 'cat_delete_sure')
            session_manager.set_user_session(message.chat.id, session, namespace="menu")
            app.send_message(message.chat.id, t(message, "confirm_delete_category"), reply_markup=markup)
        except Exception as e:
            print(traceback.format_exc())
        

############################  ADD PRODUCT  ############################

# ایجاد slug یکتا
def generate_unique_slug(model, name, max_length=50):
    from django.utils.text import slugify
    from django.utils.crypto import get_random_string
    import re
    
    if not name or not isinstance(name, str):
        # تولید slug تصادفی
        return f"p-{get_random_string(10).lower()}"
    
    # حذف کاراکترهای خاص و تولید slug
    slug = slugify(name, allow_unicode=False)
    
    # اگر slug خالی شد
    if not slug:
        # استفاده از حروف انگلیسی نام یا تولید تصادفی
        # استخراج حروف انگلیسی از نام
        english_chars = re.sub(r'[^a-zA-Z0-9]', '', name)
        if english_chars:
            slug = english_chars.lower()[:20]
        else:
            slug = f"p-{get_random_string(10).lower()}"
    
    # کوتاه کردن slug اگر طولانی باشد
    if len(slug) > max_length:
        slug = slug[:max_length-3]  # جای برای شماره می‌ماند
    
    unique_slug = slug
    counter = 1
    
    # چک کردن تکراری نبودن
    while model.objects.filter(slug=unique_slug).exists():
        unique_slug = f"{slug}-{counter}"
        counter += 1
        # اگر شماره هم اضافه کردیم و باز هم طولانی شد
        if len(unique_slug) > max_length:
            # کوتاه کردن بیشتر
            slug = slug[:max_length-5]
            unique_slug = f"{slug}-{counter}"
    
    return unique_slug

import os
import traceback
from django.conf import settings

def download_and_save_image(file_id, bot):
    try:
        # دانلود فایل
        file_info = bot.get_file(file_id)
        downloaded_file = bot.download_file(file_info.file_path)

        # مسیر ذخیره‌سازی
        save_dir = os.path.join(settings.MEDIA_ROOT, "product_images")
        os.makedirs(save_dir, exist_ok=True)  # ایجاد مسیر در صورت عدم وجود

        file_name = file_info.file_path.split('/')[-1]  # نام فایل از فایل‌پث استخراج می‌شود
        file_path = os.path.join(save_dir, file_name)

        # ذخیره فایل در سیستم
        with open(file_path, 'wb') as new_file:
            new_file.write(downloaded_file)

        # بازگشت به مسیر نسبی برای Django
        relative_file_path = os.path.join("product_images", file_name)

        return relative_file_path  # مسیر نسبی برای استفاده در قالب
    except Exception as e:
        print(f"خطا در ذخیره تصویر: {traceback.format_exc()}")
        return None



class ProductBot:
    def __init__(self, bot: TeleBot):
        self.bot = bot

    def get_name(self, message: Message):
        # ذخیره نام در Redis
        session_manager.set_user_session(message.chat.id, {"brand": False, "name_d": message.text, "price": True}, namespace="add_product")
        session = session_manager.get_user_session(message.chat.id, namespace="add_product")
        session["brand"] = False
        session["name_d"] =  message.text
        session_manager.set_user_session(message.chat.id, session, namespace="add_product")

        # ارسال منو برای وارد کردن برند
        markup = send_menu(message, [t(message, "no_brand")], message.text, [t(message, "cancel_action")])
        self.bot.send_message(message.chat.id, t(message, "enter_product_brand"), reply_markup=markup)


    def get_brand(self, message: Message):
        # ذخیره برند در Redis
        session = session_manager.get_user_session(message.chat.id, namespace="add_product")
        session["price"] = False
        session["discount"] = True
        session["brand_d"] =  None if message.text == t(message, "no_brand") else message.text

        session_manager.set_user_session(message.chat.id, session, namespace="add_product")

        # ارسال منو برای وارد کردن قیمت
        markup = send_menu(message, [t(message, "cancel_action")], message.text)
        self.bot.send_message(message.chat.id, t(message, "enter_product_price"), reply_markup=markup)



    def get_price(self, message: Message):
        try:
            # تلاش برای تبدیل پیام به عدد
            price = float(message.text)

            # بررسی اینکه قیمت معتبر است یا خیر
            if price < 10000:
                self.bot.send_message(
                    message.chat.id, t(message, "final_price_too_low"))
                return  # خروج از تابع تا کاربر دوباره قیمت وارد کند

            # ذخیره قیمت در Redis
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            session["discount"] = False
            session["status"] = True
            session["price_d"] = price

            session_manager.set_user_session(message.chat.id, session, namespace="add_product")


            # ارسال پیام برای درخواست درصد تخفیف
            self.bot.send_message(message.chat.id, t(message, "enter_discount"))
        except ValueError:
            # مدیریت خطای تبدیل مقدار نامعتبر
            self.bot.send_message(message.chat.id, t(message, "price_not_number"))



    def get_discount(self, message: Message):
        try:
            # تلاش برای تبدیل تخفیف به عدد
            discount = float(message.text)
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")

            # دریافت قیمت و محاسبه قیمت نهایی از Redis
            price = session["price_d"] or 0
            final_price = price - ((price * discount) / 100)
            # بررسی اینکه قیمت نهایی معتبر است یا خیر
            if final_price < 10000:
                self.bot.send_message(
                    message.chat.id, t(message, "final_price_too_low"))
                session["price"] = True
                session ["status"] = False
                
                session_manager.set_user_session(message.chat.id, session, namespace="add_product")
                return

            # ذخیره تخفیف در Redis و ادامه به مرحله بعد
            session["status"] = False
            session["category"] = True
            session["discount_d"] = discount

            session_manager.set_user_session(message.chat.id, session, namespace="add_product")


            # ارسال پیام برای دریافت توضیحات
            markup = send_menu(message, [t(message, "active_adj"), t(message, "inactive_adj")], message.text, [t(message, "cancel_action")])
            app.send_message(message.chat.id, t(message, "enter_status"), reply_markup=markup)
        except ValueError:
            # مدیریت خطای تبدیل مقدار نامعتبر
            self.bot.send_message(
                message.chat.id,
                t(message, "discount_not_number")
            )




    def get_stock(self, message: Message):
        try:
            # ذخیره موجودی در Redis
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            session["ask_variant_decision"] = False
            session["no_variant"] = True
            session["get_description"] = True
            session_manager.set_user_session(message.chat.id, session, namespace="add_product")

            markup = send_menu(message, [], message.text, [t(message, "cancel_action")])
            self.bot.send_message(message.chat.id, t(message, "enter_stock"), reply_markup=markup)
            
        except Exception as e:
            print(traceback.format_exc())



    def get_status(self, message: Message):
        try:
            status = message.text.strip() == t(message, "active_adj")
            session2 = session_manager.get_user_session(message.chat.id, namespace="add_product")
            session2["category"] = False
            session2["status_d"] = status
            session_manager.set_user_session(message.chat.id, session2, namespace="add_product")

            session = session_manager.get_user_session(message.chat.id, namespace="menu")
            session['product_cat_selection'] = True
            session['add_product'] = True
            # نمایش منوی دسته‌بندی اصلی
            cat = CategoryClass()
            cat.handle_category(message)
            session['add_product'] = False
            session_manager.set_user_session(message.chat.id, session, namespace="menu")


        except Exception as e:
            self.bot.send_message(message.chat.id, "خطایی رخ داده است. لطفاً دوباره تلاش کنید.")
            print(f"Error: {traceback.format_exc()}")



    def get_category(self, message: Message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            session["ask_variant_decision"] = True
                

            selected_category = Category.objects.get(title__iexact=message.text.strip(), status=True)
            print(selected_category.id)
            session["category_id"] = selected_category.id
            
            session_manager.set_user_session(message.chat.id, session, namespace="add_product")

            markup = send_menu(message, [t(message, "accurate_inventory"), t(message, "not_necessary")], "main menu", [t(message, "cancel_action")])
            self.bot.send_message(message.chat.id, t(message, "ask_variant_decision"), reply_markup=markup)
        except Category.DoesNotExist:
            self.bot.send_message(message.chat.id, t(message, "invalid_selected_category"))


    # ----------------------------
    # Variant workflow handlers
    # ----------------------------
    def get_variant_decision(self, message: Message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            session["variants"] = {}
            session["ask_variant_decision"] = False
            session["variantkey"] = True
            session_manager.set_user_session(message.chat.id, session, namespace="add_product")
            markup = send_menu(message, [t(message, "cancel_action")], message.text)
            self.bot.send_message(message.chat.id, t(message, "enter_variant_key"), reply_markup=markup)
        except Exception as e:
            print(traceback.format_exc())

    def get_variant_key(self, message: Message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            key = message.text.strip()
            variants = session.get("variants")
            if key in variants:
                self.bot.send_message(message.chat.id, t(message, "variant_key_exists"))
                return
            variants[key] = []
            session["variants"] = variants
            session["current_variant_key"] = key

            session["variantvalue"] = True
            session["variantkey"] = False
            session_manager.set_user_session(message.chat.id, session, namespace="add_product")
            self.bot.send_message(message.chat.id, t(message, "enter_variant_values", keyname=key))
        except Exception as e:
            print(traceback.format_exc())

    def get_variant_values(self, message: Message):
        try:
            # تبدیل همه ویرگول‌ها به کامای انگلیسی
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            key = message.text.strip()
            normalized_text = (
                message.text
                .replace("،", ",")   # فارسی و عربی
                .replace("，", ",")   # چینی
                .replace("﹐", ",")   # کامای عربی-قدیمی
                .replace("､", ",")   # کامای ژاپنی
            )

            # جدا کردن مقادیر
            values = [v.strip() for v in normalized_text.split(",") if v.strip()]

            key = session.get("current_variant_key")
            variants = session.get("variants")
            variants[key] = values
            session["variants"] = variants

            # رفتن به مرحله‌ی بعد
            session["variantvalue"] = False
            session["variant_add_key_answer"] = True
            session_manager.set_user_session(message.chat.id, session, namespace="add_product")

            markup = send_menu(
                message,
                [t(message, "yes"), t(message, "no")],
                "main menu",
                [t(message, "cancel_action")]
            )
            self.bot.send_message(
                message.chat.id,
                t(message, "add_another_variant_key"),
                reply_markup=markup
            )

        except Exception as e:
            print(traceback.format_exc())

    def get_variant_add_key_answer(self, message: Message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            chat_id = message.chat.id
            text = message.text.strip()

            if text == t(message, "yes"):
                session["variantkey"] = True
                session["variant_add_key_answer"] = False
                session_manager.set_user_session(message.chat.id, session, namespace="add_product")
                markup = send_menu(message, [t(message, "cancel_action")], message.text)
                self.bot.send_message(chat_id, t(message, "enter_next_variant_key"), reply_markup=markup)
                return
            if text == t(message, "no"):
                # یعنی گفته "خیر" → همه‌ی کلیدها جمع شدند
                variants = session.get("variants")
                keys = list(variants.keys())
                values = list(variants.values())

                # تمام ترکیب‌ها از همه مقدارها ساخته می‌شوند
                combinations = list(itertools.product(*values))
                session["variant_combinations"] = combinations
                session["variant_stock_index"] = 0
                session["variants_stock"] = []
                session["variants_stock_values"] = True
                session["variant_add_key_answer"] = False

                session_manager.set_user_session(message.chat.id, session, namespace="add_product")

                # پیام مقدمه
                readable_keys = "، ".join(keys)
                markup = send_menu(message, [t(message, "cancel_action")], message.text)
                self.bot.send_message(
                    chat_id,
                    t(message, "variant_stock_intro", keys=readable_keys),
                    reply_markup=markup
                )

                # ترکیب اول را آماده کنیم
                first_combo = combinations[0]
                combo_text = " ".join([f"{keys[i]}: {first_combo[i]}" for i in range(len(keys))])

                self.bot.send_message(chat_id, t(message, "enter_variant_stock", combo=combo_text))
            else:
                markup = send_menu(
                    message,
                    [t(message, "yes"), t(message, "no")],
                    "main menu",
                    [t(message, "cancel_action")]
                )
                self.bot.send_message(
                    message.chat.id,
                    t(message, "choose_from_options"),
                    reply_markup=markup
                )
                
        except Exception as e:
            print(traceback.format_exc())



   
    def get_variants_stock_values(self, message: Message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            chat_id = message.chat.id
            stock_str = message.text.strip()
    
            # --- اعتبارسنجی اولیه
            if not stock_str.isdigit():
                self.bot.send_message(chat_id, t(message, "invalid_stock_input"))
                return
    
            stock = int(stock_str)
    
            # --- بازیابی داده‌ها
            combinations = session.get("variant_combinations", [])
            index = session.get("variant_stock_index", 0)
            variants_stock = session.get("variants_stock", [])
            variants = session.get("variants", {})
            keys = list(variants.keys())
    
            # --- ذخیره موجودی ترکیب فعلی
            variants_stock.append({
                "combination": combinations[index],
                "stock": stock
            })
            session["variants_stock"] = variants_stock
    
            # --- محاسبه مجموع موجودی‌ها
            total_stock = sum(item["stock"] for item in variants_stock)
            session["get_stock_d"] = total_stock  # ذخیره موجودی کل
    
            # --- رفتن به ترکیب بعدی
            index += 1
    
            # بررسی آیا این آخرین واریانت بود؟
            is_last_variant = (index >= len(combinations))
    
            if index < len(combinations):
                # هنوز ترکیب‌های باقی مانده وجود دارد
                session["variant_stock_index"] = index
                session_manager.set_user_session(chat_id, session, namespace="add_product")
    
                # نمایش ترکیب بعدی برای کاربر
                combo_text = " ".join([f"{keys[i]}: {combinations[index][i]}" for i in range(len(keys))])
                print(f"combo_text: {combo_text}")
                session_manager.set_user_session(message.chat.id, session, namespace="add_product")
                print(t(message, "variant_stock_question", combo_text=combo_text))
                self.bot.send_message(chat_id, t(message, "variant_stock_question", combo_text=combo_text))
            else:
                # تمام ترکیب‌ها تکمیل شده‌اند
                session["variants_stock_values"] = False
                session["get_description"] = True
    
                # نمایش خلاصه و درخواست توضیحات
                total = sum(item["stock"] for item in variants_stock)
                text = t(message, "variant_stock_saved", total=total) + "\n\n" + t(message, "enter_description")
                
                markup = send_menu(
                    message,
                    [t(message, "no_description")],
                    "main menu",
                    [t(message, "cancel_action")]
                )
                self.bot.send_message(chat_id, text, reply_markup=markup)


                session_manager.set_user_session(chat_id, session, namespace="add_product")
    
                # فقط وقتی آخرین واریانت را گرفتیم "سلام دنیا" پرینت کن
                if is_last_variant:
                    print("سلام دنیا")
                    
                    session["get_description"] = False
                    session["get_attribute"] = True

                    session_manager.set_user_session(chat_id, session, namespace="add_product")
    
        except Exception as e:
            print(f"Error in get_variants_stock_values: {e}\n{traceback.format_exc()}")
            self.bot.send_message(chat_id, "خطایی رخ داده است. لطفاً دوباره تلاش کنید.")
    
 

    def get_description(self, message: Message):
        # ذخیره خصوصیات محصول در Redis
        description = None if message.text == t(message, "no_description") else message.text
        session = session_manager.get_user_session(message.chat.id, namespace="add_product")
        session["get_attribute"] = False
        session["get_more_attributes"] = True
        session["get_description_d"] = description

        # تغییر وضعیت به مرحله ویژگی‌ها
        session_manager.set_user_session(message.chat.id, session, namespace="add_product")

        # ایجاد دکمه پایان

        
        markup = send_menu(message, [t(message, "cancel_action")], message.text)
        fake_message = self.bot.send_message(message.chat.id, "انبارداری موفق", reply_markup=markup)
        self.bot.delete_message(message.chat.id, fake_message.message_id)

        markup = types.InlineKeyboardMarkup()
        finish_button = types.InlineKeyboardButton(text=t(message, "no_ads_features"), callback_data="finish_attributes")
        markup.add(finish_button)

        # ارسال پیام برای درخواست ویژگی‌های تبلیغاتی
        self.bot.send_message(
            message.chat.id,
            t(message, "enter_ads_features"),
            reply_markup=markup
        )

        # ارسال منو اصلی
        markup = send_menu(message, [], "main menu", [t(message, "cancel_action")])
        




    def get_product_attributes(self, message: Message):
        try:
            # نمایش پیام برای وارد کردن ویژگی‌های محصول
            markup = types.InlineKeyboardMarkup()
            finish_button = types.InlineKeyboardButton(text=t(message, "finish"), callback_data="finish_attributes")
            markup.add(finish_button)

            key = message.text.split(":")[0]  # کلید ویژگی (مانند "وزن")
            if ":" not in message.text:
                value = ""
            else:
                value = message.text.split(":")[1]  # مقدار ویژگی (مانند "1kg")

            # بازیابی ویژگی‌ها از Redis
            session = session_manager.get_user_session(message.chat.id, namespace="add_product")
            if "product_attributes" not in session:
                session["product_attributes"] = {}
            session["product_attributes"][key] = value


            # ذخیره ویژگی‌های جدید در Redis
            session_manager.set_user_session(message.chat.id, session, namespace="add_product")

            self.bot.send_message(
                message.chat.id,
                t(message, "enter_ads_features"),
                reply_markup=markup
            )
        except Exception as e:
            self.bot.send_message(message.chat.id, "خطا در ذخیره ویژگی رخ داده است.")
            print(f"Error: {traceback.format_exc()}")

    def handle_finish_attributes(self, callback_query: types.CallbackQuery):
        try:
            chat_id = callback_query.message.chat.id
            session = session_manager.get_user_session(callback_query.message.chat.id, namespace="add_product")
            print(session["category_id"])
            session["get_main_image"] = False
            session["get_additional_images"] = True
            session["get_more_attributes"] = False
            session_manager.set_user_session(callback_query.message.chat.id, session, namespace="add_product")


            # تغییر وضعیت به تصویر اصلی
            self.bot.send_message(chat_id, t(callback_query.message, "send_main_image"))
        except Exception as e:
            self.bot.send_message(callback_query.message.chat.id, "خطا در ذخیره ویژگی رخ داده است.")
            print(f"Error: {traceback.format_exc()}")

    def register_handle_finish_attributes(self):
        self.bot.callback_query_handler(func=lambda call: call.data == 'finish_attributes')(self.handle_finish_attributes)



    def get_main_image(self, message: Message):
        try:
            # بررسی اینکه آیا کاربر عکس فرستاده یا نه
            if not message.photo:
                self.bot.send_message(message.chat.id, t(message, "please_send_image"))
                # ثبت مجدد هندلر برای دریافت عکس
                self.bot.register_next_step_handler(message, self.get_main_image)
                return

            # دانلود و ذخیره تصویر
            file_id = message.photo[-1].file_id
            saved_path = download_and_save_image(file_id, self.bot)

            if saved_path:
                # ذخیره مسیر تصویر در Redis
                session = session_manager.get_user_session(message.chat.id, namespace="add_product")
                session["get_additional_images"] = False
                session["process_accomplished"] = True
                session["main_image"] = saved_path

                # تغییر وضعیت به مرحله بعدی (تصاویر اضافی)
                session_manager.set_user_session(message.chat.id, session, namespace="add_product")
                
                self.bot.send_message(message.chat.id, t(message, "send_three_more_images"))
            else:
                self.bot.send_message(message.chat.id, "خطا در ذخیره تصویر اصلی رخ داده است.")
                # اگر خطا در ذخیره عکس اتفاق افتاد، دوباره درخواست عکس بده
                self.bot.register_next_step_handler(message, self.get_main_image)
        except Exception as e:
            self.bot.send_message(message.chat.id, "خطایی رخ داده است. لطفاً دوباره تلاش کنید.")
            print(f"Error: {traceback.format_exc()}")
            # در صورت خطا، دوباره درخواست عکس بده
            self.bot.register_next_step_handler(message, self.get_main_image)



    def generate_readable_sku(self, product_name, combo_list, product_id, attempt=1):
        """
        Generate a readable, unique SKU for each product variant
        Example output: SHIRT-BLUE-L-001
        """
        base_name = re.sub(r'[^A-Z0-9]', '', product_name.upper().replace(" ", ""))[:6]
        variant_part = "-".join(str(v).upper().replace(" ", "")[:4] for v in combo_list)
        sku = f"{base_name}-{variant_part}-{attempt:03d}"

        if ProductVariant.objects.filter(sku=sku).exists():
            return self.generate_readable_sku(product_name, combo_list, product_id, attempt + 1)
        return sku


    def get_additional_images(self, message: Message):
        try:
            chat_id = message.chat.id
            session = session_manager.get_user_session(chat_id, namespace="add_product")
            additional_images = session.get("additional_images", [])
            file_id = message.photo[-1].file_id
            saved_image = download_and_save_image(file_id, self.bot)

            if not saved_image:
                self.bot.send_message(chat_id, t(message, "extra_image_save_failed"))
                return 
            
            additional_images.append(saved_image)
            session["additional_images"] = additional_images
            
            # تعداد عکس‌های مورد نیاز (می‌تواند دینامیک باشد)
            required_images = 3  # یا می‌توانی از session بخوانی: session.get("required_images", 3)
            
            # بررسی آیا یک عکس به آخرین عکس مانده؟
            current_count = len(additional_images)
            remaining = required_images - current_count
            
            if remaining == 1:  # یک عکس به آخرین عکس مانده
                session["process_getout"] = True
                session["process_accomplished"] = False
            
            if current_count < required_images:
                markup = send_menu(message, [t(message, "cancel_action")], message.text)
                self.bot.send_message(chat_id, t(message, "send_extra_images", pic_num=remaining), reply_markup=markup)
                session_manager.set_user_session(chat_id, session, namespace="add_product")
                return

            # اگر به تعداد مورد نیاز رسیدیم، ادامه پردازش...
            # Retrieve user, store, and category
            profile = ProfileModel.objects.get(tel_id=message.from_user.id)
            store = Store.objects.get(owner=profile)

            category_id = session.get("category_id")
            if not category_id:
                self.bot.send_message(chat_id, t(message, "invalid_selected_category"))
                return
            category = Category.objects.get(id=category_id, status=True)

            # Product basic info
            name = session.get("name_d")
            slug = generate_unique_slug(Product, name)
            print(f"DEBUG: name='{name}', type={type(name)}")
            print(f"DEBUG: Generated slug for '{name}': '{slug}'")  # اضافه کردن این خط
            
            # یا لاگ کامل
            import logging
            logger = logging.getLogger(__name__)
            logger.debug(f"Generating slug for product: name='{name}', slug='{slug}'")
            
            price = session.get("price_d")
            discount = session.get("discount_d")
            stock = session.get("get_stock_d")
            status = session.get("status_d")
            brand = session.get("brand_d")
            description = session.get("get_description_d")
            main_image = session.get("main_image")

            product = Product.objects.create(
                profile=profile,
                store=store,
                name=name,
                brand=brand,
                price=price,
                discount=discount,
                stock=stock,
                status=status,
                category=category,
                description=description,
                main_image=main_image,
                slug=slug
            )

            # Product attributes
            product_attrs = session.get("product_attributes", {})
            for key, value in product_attrs.items():
                ProductAttribute.objects.create(product=product, key=key, value=value)

            # Variants
            variants = session.get("variants", {})
            variant_combinations = session.get("variant_combinations", [])
            variants_stock = session.get("variants_stock", [])
            entries = variants_stock or [{"combination": combo, "stock": 0} for combo in variant_combinations]

            option_cache = {}
            value_cache = {}
            keys_order = list(variants.keys())

            if entries:
                for entry in entries:
                    combo_raw = entry.get("combination") or entry.get("combo")
                    stock = int(entry.get("stock", 0)) if entry.get("stock") else 0
                    combo_list = list(combo_raw.values()) if isinstance(combo_raw, dict) else list(combo_raw)

                    if len(combo_list) != len(keys_order):
                        print(f"[warn] combo length mismatch: keys={keys_order} combo={combo_list}")
                        continue

                    sku = self.generate_readable_sku(product.name, combo_list, product.id)
                    variant = ProductVariant.objects.create(product=product, stock=stock, sku=sku)

                    for i, key_name in enumerate(keys_order):
                        val = combo_list[i]

                        if key_name not in option_cache:
                            option_obj, _ = ProductOption.objects.get_or_create(product=product, name=key_name)
                            option_cache[key_name] = option_obj
                        else:
                            option_obj = option_cache[key_name]

                        if (key_name, val) not in value_cache:
                            value_obj, _ = ProductOptionValue.objects.get_or_create(option=option_obj, value=val)
                            value_cache[(key_name, val)] = value_obj
                        else:
                            value_obj = value_cache[(key_name, val)]

                        variant.values.add(value_obj)
            else:
                # If product has no variants, generate a simple SKU
                ProductVariant.objects.create(
                    product=product,
                    stock=stock,
                    sku=f"{re.sub(r'[^A-Z0-9]', '', product.name.upper())[:8]}-{product.id:03d}"
                )

            # Save extra images
            for image_path in additional_images:
                ProductImage.objects.create(product=product, image=image_path)


            session["code"] = product.code
            session_manager.set_user_session(chat_id, session, namespace="add_product")

            # Send success message
            self.bot.send_message(chat_id, t(message, "product_saved"))
            

        except Exception as e:
            print(f"Error in get_additional_images: {e}\n{traceback.format_exc()}")
            self.bot.send_message(message.chat.id, t(message, "product_save_failed"))


    def delete(self, message: Message):
        try:
            if message.text == t(message, "cancel_action"):
                self.cancle_request(message)
            else:
                code = message.text
                try:
                    product = Product.objects.get(code=code)
                except Product.DoesNotExist:
                    self.bot.send_message(message.chat.id, t(message, "product_code_not_found"))
                    return
                try:
                    product = Product.objects.get(code=code, store__owner__tel_id=message.chat.id)
                except Product.DoesNotExist:
                    self.bot.send_message(message.chat.id, t(message, "cannot_remove_other_store_item"))
                    return

                attributes = product.attributes.all()
                # ارسال پیام محصول به کاربر
                producthandler = ProductHandler(app=self.bot, product=product, current_site=settings_current_site, attributes=attributes, chat_id=message.chat.id)
                producthandler.send_product_message(chat_id=message.chat.id, buttons=False)

                # ذخیره اطلاعات محصول در Redis
                session = session_manager.get_user_session(message.chat.id, namespace="delete_product")
                session['enter_product_code_to_delete'] = False
                session['delete_product_confirm'] = True
                session['code'] = code
                session_manager.set_user_session(message.chat.id, session, namespace="delete_product")

                menu = [t(message, "yes_im_sure"), t(message, "cancel_action")]
                markup = send_menu(message, menu, "main menu", home_menu)
                self.bot.send_message(message.chat.id, t(message, "confirm_delete_product"), reply_markup=markup)

        except Exception as e:
            error_message = traceback.format_exc()  # دریافت Traceback کامل
            print(f"Error in handle_buttons: {e}\n{error_message}")




    def delete_confirm(self, message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="delete_product")

            # بازیابی کد محصول از Redis
            product_code = session['code']
            if product_code:
                try:
                    product = Product.objects.get(code=product_code, store__owner__tel_id=message.chat.id)
                    if message.text == t(message, "yes_im_sure"):
                        # حذف محصول از دیتابیس
                        product.delete()
                        
                        # ارسال پیام موفقیت‌آمیز به کاربر
                        self.bot.send_message(message.chat.id, t(message, "product_deleted"))

                    elif message.text == t(message, "cancel_action"):
                        self.cancle_request(message)
                        return

                    session_manager.reset_user_session(message.chat.id, namespace="delete_product")
                except Product.DoesNotExist:
                    self.bot.send_message(message.chat.id, t(message, "cannot_remove_other_store_item"))
                    return
            else:
                self.bot.send_message(message.chat.id, "کد محصول ذخیره نشده است.")
                return

        except Exception as e:
            self.bot.send_message(message.chat.id, "خطایی رخ داده است. لطفاً دوباره تلاش کنید.")
            print(f"Error: {traceback.format_exc()}")

    def deactivate(self, message):
        try:
            code = message.text
            try:
                product = Product.objects.get(code=code)
            except Product.DoesNotExist:
                self.bot.send_message(message.chat.id, t(message, "product_code_not_found"))
                return
            try:
                product = Product.objects.get(code=code, store__owner__tel_id=message.chat.id)
            except Product.DoesNotExist:
                self.bot.send_message(message.chat.id, t(message, "cannot_disable_other_store_item"))
                return

            attributes = product.attributes.all()
            # ارسال پیام محصول به کاربر
            producthandler = ProductHandler(app=self.bot, product=product, current_site=settings_current_site, attributes=attributes, chat_id=message.chat.id)
            producthandler.send_product_message(chat_id=message.chat.id, buttons=False)

            # ذخیره اطلاعات محصول در Redis
            state_manager = RedisStateManager(message.chat.id)
            state_manager.save_user_data("product_code", code)
            session = {'deactivate_product_confirm': True, 'code': code}
            session_manager.set_user_session(message.chat.id, session, namespace="deactivate_product")
            session = session_manager.get_user_session(message.chat.id, namespace="menu")
            session["deavtivate_product"] = False
            session["deactivate_product_confirm"] = True
            session_manager.set_user_session(message.chat.id, session, namespace="menu")

            menu = [t(message, "yes_im_sure"), t(message, "cancel_action")]
            markup = send_menu(message, menu, "main menu", home_menu)
            action_text = t(message, "deactivate") if product.status else t(message, "activate")
            msg_text = t(message, "confirm_toggle_product", action=action_text)
            self.bot.send_message(message.chat.id, msg_text, reply_markup=markup)

        except Exception as e:
            error_message = traceback.format_exc()  # دریافت Traceback کامل
            print(f"Error in handle_buttons: {e}\n{error_message}")

    def deactivate_confirm(self, message):
        try:
            session = session_manager.get_user_session(message.chat.id, namespace="deactivate_product")
            # بازیابی کد محصول از Redis
            product_code = session.get("code")
            if product_code:
                try:
                    product = Product.objects.get(code=product_code, store__owner__tel_id=message.chat.id)
                    if message.text == t(message, "yes_im_sure"):
                        # حذف محصول از دیتابیس
                        if product.status:
                            product.status = False
                            product.save()
                        else:
                            product.status = True
                            product.save()

                        # ارسال پیام موفقیت‌آمیز به کاربر
                        action_text = t(message, "activated") if product.status else t(message, "deactivated")
                        self.bot.send_message(message.chat.id, t(message, "product_toggled", action=action_text))

                    elif message.text == t(message, "cancel_action"):
                        self.cancle_request(message)
                        return

                except Product.DoesNotExist:
                    self.bot.send_message(message.chat.id, t(message, "cannot_disable_other_store_item"))
                    return
            else:
                self.bot.send_message(message.chat.id, "کد محصول ذخیره نشده است.")
                return
            
        except Exception as e:
            self.bot.send_message(message.chat.id, "خطایی رخ داده است. لطفاً دوباره تلاش کنید.")
            print(f"Error: {traceback.format_exc()}")

    def cancle_request(self, message):
        try:
            if subscription.subscription_offer(message):
                # بازیابی منوها از Redis
                session_manager.reset_user_session(message.chat.id, namespace="add_product")
                return
        except Exception as e:
            self.bot.send_message(message.chat.id, "خطایی رخ داده است. لطفاً دوباره تلاش کنید.")
            print(f"Error: {traceback.format_exc()}")



############################  SEND PAYMENT LINK  ############################

def send_payment_link(app, context):
    chat_id = update.message.chat_id
    email = "example@test.com"  # ایمیل کاربر
    mobile = ProfileModel.objects.get(tel_id=chat_id).phone  # شماره موبایل کاربر
    amount = 100000  # مبلغ پرداخت
    description = "توضیحات کالا"

    # ساخت لینک پرداخت
    payment_url = f"http://intelleum.ir:8443/buy/{amount}/{description}/?email={email}&mobile={mobile}"

    return payment_url

############################  SEND PRODUCT MESSAGE  ############################

from telethon import Button
import traceback
from telethon.tl.types import InputMediaPhoto
from django.db import models
from django.core.cache import cache
import threading
from asgiref.sync import sync_to_async

# @add_performance_monitoring_to_class
class ProductHandler:
    """مدیریت ارسال پیام و اطلاعات محصول - نسخه بهینه‌شده"""
    
    def __init__(self, app, product, current_site, photos=None, attributes=None, chat_id=None):
        self.app = app
        self.product = product
        self.current_site = settings_current_site
        self.user_manager = UserOrderManager()
        self.photos = photos or []
        self.attributes = attributes
        self._variants_data_cache = None
        self.chat_id = chat_id

    def get_product_variants_data(self):
        """دریافت همه داده‌های واریانت در یک کوئری با کشینگ"""
        if self._variants_data_cache:
            return self._variants_data_cache
            
        cache_key = f"product_{self.product.id}_full_variants_data"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            self._variants_data_cache = cached_data
            return cached_data
        
        # بارگذاری بهینه با prefetch_related
        from products.models import ProductVariant
        
        variants = ProductVariant.objects.filter(
            product_id=self.product.id
        ).prefetch_related('values__option')
        
        variants_dict = {}
        variants_list = []
        
        for variant in variants:
            variant_data = {
                'id': variant.id,
                'sku': variant.sku,
                'stock': variant.stock,
                'price_override': variant.price_override,
                'values': {}
            }
            
            for option_value in variant.values.all():
                key = option_value.option.name.capitalize()
                value = option_value.value
                variant_data['values'][key] = value
                
                if key not in variants_dict:
                    variants_dict[key] = set()
                variants_dict[key].add(value)
            
            variants_list.append(variant_data)
        
        def sort_variant_values(values):
            numeric_values = []
            string_values = []
            
            for value in values:
                try:
                    numeric_values.append(float(value) if '.' in str(value) else int(value))
                except (ValueError, TypeError):
                    string_values.append(str(value))
            
            sorted_numeric = sorted(numeric_values)
            sorted_strings = sorted(string_values)
            
            return [str(val) for val in sorted_numeric] + sorted_strings

        result = {
            'variants_dict': {key: sort_variant_values(values) for key, values in variants_dict.items()},
            'variants_list': variants_list
        }
        
        cache.set(cache_key, result, timeout=300)
        self._variants_data_cache = result
        
        return result

    async def async_get_product_variants_data(self):
        variants = await sync_to_async(list)(
            ProductAttribute.objects.filter(product=self.product)
        )

        variants_dict = {}

        for variant in variants:
            key = variant.key
            value = variant.value
            if key not in variants_dict:
                variants_dict[key] = []
            if value:
                variants_dict[key].append(value)

        return {
            "variants_dict": variants_dict
        }

    def get_variants_dict(self, variants=None):
        """تبدیل واریانت‌ها به دیکشنری - کاملاً از کش استفاده می‌کند"""
        variants_data = self.get_product_variants_data()
        return variants_data['variants_dict']

    def get_variant_by_selected_values(self, product, selected_values):
        """پیدا کردن واریانت بر اساس مقادیر انتخاب شده"""
        variants_data = self.get_product_variants_data()
        
        print(f"🔍 [VARIANT DEBUG] Looking for variant with: {selected_values}")
        print(f"🔍 [VARIANT DEBUG] Total variants: {len(variants_data['variants_list'])}")
        
        for variant_data in variants_data['variants_list']:
            variant_values = variant_data['values']
            
            # بررسی تطابق کامل
            match = True
            for key, selected_value in selected_values.items():
                variant_value = variant_values.get(key)
                
                if not variant_value:
                    match = False
                    break
                    
                selected_clean = str(selected_value).strip().lower()
                variant_clean = str(variant_value).strip().lower()
                
                if selected_clean != variant_clean:
                    match = False
                    break
            
            if match:
                print(f"✅ [VARIANT DEBUG] EXACT MATCH FOUND: {variant_data['id']}")
                from products.models import ProductVariant
                return ProductVariant.objects.get(id=variant_data['id'])
        
        print("❌ [VARIANT DEBUG] No exact matching variant found")
        return None

    def format_price(self):
        """فرمت‌بندی قیمت بدون کوئری دیتابیس"""
        formatted_price = "{:,.0f}".format(float(self.product.price))
        formatted_final_price = "{:,.0f}".format(float(self.product.final_price))

        if self.product.discount > 0:
            return (
                f"🏃 {self.product.discount} % تخفیف\n"
                f"💵 {t('message', 'price', chat_id=self.chat_id)}: <s>{formatted_price}</s> تومان ⬅ {formatted_final_price} تومان"
            )
        return f"💵 {t('message', 'price', chat_id=self.chat_id)}: {formatted_price} تومان"

    
    def build_attributes_text(self):
        if not self.attributes:
            return ""
        return "\n".join(
            [f"✨ {a.key}: {a.value}" if a.value else f"✨ {a.key}" for a in self.attributes]
        ) + "\n\n"


    def build_variants_text(self, variants_dict):
        if not variants_dict:
            return ""
        lines = [
            f"✅ {key}: {', '.join(values)}"
            for key, values in variants_dict.items()
        ]
        return "\n".join(lines) + "\n\n"


    async def async_get_product_variants_data(self):
        variants = await sync_to_async(list)(
            self.product.variants.prefetch_related("values__option").all()
        )

        variants_dict = defaultdict(set)

        for variant in variants:
            values = await sync_to_async(list)(variant.values.all())
            for val in values:
                variants_dict[val.option.name].add(val.value)

        return {
            "variants_dict": {
                k: sorted(list(v)) for k, v in variants_dict.items()
            }
        }



    
    def generate_caption(self):
        brand_text = f"🔖 {t("message", "product_brand", chat_id=self.chat_id)}: {self.product.brand}\n" if self.product.brand else ""
        description_text = f"{self.product.description}\n" if self.product.description else ""

        attribute_text = self.build_attributes_text()

        variants_data = self.get_product_variants_data()
        variants_text = self.build_variants_text(variants_data.get("variants_dict"))

        
        product_name = t("message", "product_name", chat_id=self.chat_id)
        product_code = t("message", "product_code", chat_id=self.chat_id)


        return (
            f"\n⭕️ {product_name}: {self.product.name}\n"
            f"{brand_text}"
            f"{product_code}: {self.product.code}\n\n"
            f"{description_text}\n"
            f"{attribute_text}"
            f"{variants_text}"
            f"{self.format_price()}\n"
        )


    async def async_generate_caption(self):
        brand_text = f"🔖 برند کالا: {self.product.brand}\n" if self.product.brand else ""
        description_text = f"{self.product.description}\n" if self.product.description else ""
    
        # Attributes
        attribute_text = ""
        if self.attributes:
            attribute_text = "\n✨ ".join(
                [f"{attr.key}: {attr.value}" if attr.value else f"{attr.key}" for attr in self.attributes]
            )
            attribute_text = f"✨ {attribute_text}\n\n"
    
        # واریانت‌ها
        variants_data = await self.async_get_product_variants_data()
    
        variants_text = ""
        if variants_data['variants_dict']:
            variant_lines = [
                f"{key}: {', '.join(values)}"
                for key, values in variants_data['variants_dict'].items()
            ]
            variants_text = "✅ " + "\n✅ ".join(variant_lines) + "\n\n"
    
        # قیمت (sync_to_async درست)
        price_text = await sync_to_async(self.format_price)()
        
        return (
            f"\n⭕️ نام کالا: {self.product.name}\n"
            f"{brand_text}"
            f"کد کالا: {self.product.code}\n\n"
            f"{description_text}\n"
            f"{attribute_text}"
            f"{variants_text}"
            f"📫 ارسال به تمام نقاط کشور\n\n"
            f"{price_text}\n"
        )




    async def send_product_channel(self, chat_id, buttons=True):
        try:
            caption = await self.async_generate_caption()

            if not self.photos:
                await self.app.send_message(
                    chat_id,
                    caption,
                    parse_mode="html",
                    buttons=[[Button.inline("🛒 خرید", b"buy_now")]] if buttons else None
                )
                return

            files = [await self.app.upload_file(p) for p in self.photos]

            await self.app.send_file(
                chat_id,
                files,
                caption=caption,
                parse_mode="html",
                supports_streaming=True
            )

            if buttons:
                await self.app.send_message(
                    chat_id,
                    "👇👇👇",
                    buttons=[[Button.inline("🛒 خرید", b"buy_now")]]
                )

        except Exception:
            print("❌ send_product_channel error:\n", traceback.format_exc())



    def send_product_message(self, chat_id, buttons=True):
        """ارسال محصول با دکمه‌های سریع"""
        from AI.settings import BASE_DIR
        try:
            photos = [
                types.InputMediaPhoto(open(os.path.join(BASE_DIR, self.product.main_image.path), 'rb'), caption=self.generate_caption(), parse_mode='HTML')
            ] + [
                types.InputMediaPhoto(open(os.path.join(BASE_DIR, i.image.path), 'rb')) for i in self.product.images.all()
            ]
            

            if len(photos) > 10:
                photos = photos[:10]

            self.app.send_media_group(chat_id, media=photos)
            
            # استفاده از نسخه سریع برای دکمه‌ها
            if buttons:
                self.send_buttons(chat_id)  # این حالا نسخه سریع است
                
        except Exception as e:
            error_message = traceback.format_exc()
            print(f"Error in send_product_message: {e}\n{error_message}")

    def send_buttons(self, chat_id):
        """ارسال سریع دکمه‌ها بدون تأخیر"""
        try:
            # محاسبه سریع داده‌های مورد نیاز
            buttons_data = self._prepare_buttons_data(chat_id)
            if not buttons_data:
                return
                
            # ارسال فوری دکمه‌ها
            self._send_buttons_from_data(chat_id, buttons_data)
            
        except Exception as e:
            print(f"Error in send_buttons: {traceback.format_exc()}")

    def _prepare_buttons_data(self, chat_id):
        """آماده‌سازی سریع داده‌های دکمه"""
        try:
            # دریافت session
            session = SessionManager()
            user_session = session.get_user_session(chat_id, namespace="variants")
            variant_states = user_session.get(str(self.product.code), {})
            
            # کوئری‌های موازی برای داده‌های ضروری
            profile = ProfileModel.objects.get(tel_id=chat_id)
            cart, _ = Cart.objects.get_or_create(profile=profile)
            
            # استفاده از داده‌های کش شده
            variants_dict = self.get_variants_dict(self.product.variants.all())
            
            # محاسبه سریع مقادیر انتخاب شده
            selected_values = {}
            for i, key in enumerate(variants_dict.keys()):
                if str(i) in variant_states:
                    values_list = list(variants_dict[key])
                    selected_index = variant_states[str(i)]
                    if selected_index < len(values_list):
                        selected_values[key] = values_list[selected_index]
            
            # پیدا کردن واریانت
            variant = None
            if selected_values:
                variant = self.get_variant_by_selected_values(self.product, selected_values)
            
            # وضعیت سبد خرید
            current_quantity = 0
            cart_item_exists = False
            
            if variant:
                cart_item = CartItem.objects.filter(
                    cart=cart, 
                    product=self.product, 
                    variant=variant
                ).first()
                if cart_item:
                    current_quantity = cart_item.quantity
                    cart_item_exists = True
            else:
                cart_item = CartItem.objects.filter(
                    cart=cart, 
                    product=self.product, 
                    variant__isnull=True
                ).first()
                if cart_item:
                    current_quantity = cart_item.quantity
                    cart_item_exists = True

            return {
                'variant_states': variant_states,
                'cart': cart,
                'variants_dict': variants_dict,
                'selected_values': selected_values,
                'variant': variant,
                'current_quantity': current_quantity,
                'cart_item_exists': cart_item_exists
            }
            
        except Exception as e:
            print(f"Error in _prepare_buttons_data: {traceback.format_exc()}")
            return None

    def _send_buttons_from_data(self, chat_id, buttons_data):
        """ارسال فوری دکمه‌ها با داده‌های از پیش محاسبه شده"""
        try:
            print(buttons_data)
            variant_states = buttons_data['variant_states']
            cart = buttons_data['cart']
            variants_dict = buttons_data['variants_dict']
            variant = buttons_data['variant']
            current_quantity = buttons_data['current_quantity']
            cart_item_exists = buttons_data['cart_item_exists']

            # ساخت سریع دکمه‌ها
            buttons = []
            handlers = {}
            
            if cart_item_exists:
                # حالت شمارنده
                variant_id = variant.id if variant else "0"
                buttons.extend([
                    ("➕", f"increase_{self.product.code}_{variant_id}", 2),
                    (f"{current_quantity}", "count", 1),
                    ("➖", f"decrease_{self.product.code}_{variant_id}", 0),
                ])
                
                handlers.update({
                    f"increase_{self.product.code}": self.handle_buttons,
                    f"decrease_{self.product.code}": self.handle_buttons,
                })
                
                # دکمه‌های واریانت
                for i, (key, values) in enumerate(variants_dict.items()):
                    current_index = variant_states.get(str(i), 0)
                    current_value = values[current_index] if current_index < len(values) else values[0]
                    
                    buttons.extend([
                        ("⏪", f"VarPrev_{self.product.code}_{i}", i * 3 + 3),
                        (f"{key}: {current_value}", f"var_{i}", i * 3 + 4),
                        ("⏩", f"VarNext_{self.product.code}_{i}", i * 3 + 5),
                    ])
                    
                    handlers.update({
                        f"VarPrev_{self.product.code}_{i}": self.handle_variant_navigation,
                        f"VarNext_{self.product.code}_{i}": self.handle_variant_navigation,
                    })

                # دکمه سبد خرید
                total_cart_items = cart.total_items()
                buttons.append((f"{t("message", "menu_cart", chat_id=chat_id)} ({total_cart_items})", "view_cart", len(buttons) + 2))
                
                # لیآوت
                if variants_dict:
                    button_layout = [3] + [3] * len(variants_dict) + [1]
                else:
                    button_layout = [3, 1]
                    
            else:
                # حالت اولیه
                buttons.extend([
                    (t("message", "add_to_cart", chat_id=chat_id), f"addtocart_{self.product.code}", 1),
                    (t("message", "comments", chat_id=chat_id), f"comments_{self.product.code}", 0),
                ])

                
                handlers.update({
                    f"addtocart_{self.product.code}": self.handle_add_to_cart,
                    f"comments_{self.product.code}": self.handle_comments,
                })
                
                button_layout = [2]

                variant = None

            # متن اطلاع‌رسانی
            stock_info = variant.stock if variant else self.product.stock
            print("_send_buttons_from_data")
            text = t("message", "order_up_to_stock", chat_id=chat_id, stock_info=stock_info)
            
            # ارسال فوری
            markup = SendMarkup(
                bot=self.app,
                chat_id=chat_id,
                text=text,
                buttons=buttons,
                button_layout=button_layout,
                handlers=handlers
            )
            markup.send()
            
        except Exception as e:
            print(f"Error in _send_buttons_from_data: {traceback.format_exc()}")

    def _send_buttons_safe(self, chat_id):
        """متد قدیمی برای سازگاری - حالا از نسخه سریع استفاده می‌کند"""
        self.send_buttons(chat_id)

    def show_initial_state(self, chat_id, message_id, product):
        """نمایش حالت اولیه"""
        try:
            buttons = [
                (t("message", "add_to_cart", chat_id=chat_id), f"addtocart_{product.code}", 1),
                (t("message", "comments", chat_id=chat_id), f"comments_{product.code}", 0),
            ]
            
            handlers = {
                f"addtocart_{product.code}": self.handle_add_to_cart,
                f"comments_{product.code}": self.handle_comments,
            }
            
            stock_info = product.stock
            print("initail state")
            text = t("message", "order_up_to_stock", chat_id=chat_id, stock_info=stock_info)

            markup = SendMarkup(
                bot=self.app,
                chat_id=chat_id,
                text=text,
                buttons=buttons,
                button_layout=[2],
                handlers=handlers,
            )
            
            markup.edit(message_id)

        except Exception as e:
            print(f"❌ Error in show_initial_state: {traceback.format_exc()}")

    def handle_add_to_cart(self, call):
        """مدیریت افزودن به سبد خرید"""
        try:
            data = call.data.split("_")
            product_code = str(data[-1])
            chat_id = call.message.chat.id
            message_id = call.message.message_id

            product = Product.objects.get(code=product_code)
            profile = ProfileModel.objects.get(tel_id=chat_id)
            cart, _ = Cart.objects.get_or_create(profile=profile)

            session = SessionManager()
            user_session = session.get_user_session(chat_id, namespace="variants")
            variant_states = user_session.get(str(product_code), {})

            variants_dict = self.get_variants_dict(product.variants.all())
            selected_values = {}
            
            for i, key in enumerate(variants_dict.keys()):
                if str(i) in variant_states:
                    values_list = list(variants_dict[key])
                    selected_index = variant_states[str(i)]
                    if selected_index < len(values_list):
                        selected_values[key] = values_list[selected_index]

            variant = None
            if selected_values:
                variant = self.get_variant_by_selected_values(product, selected_values)

            cart_item = None
            if variant:
                cart_item = CartItem.objects.filter(cart=cart, product=product, variant=variant).first()
            else:
                if not product.variants.exists():
                    cart_item = CartItem.objects.filter(cart=cart, product=product, variant__isnull=True).first()

            if not cart_item:
                cart_item = CartItem.objects.create(
                    cart=cart, 
                    product=product,
                    variant=variant,
                    quantity=0
                )

            self.update_product_message(chat_id, message_id, product, cart)
            # self.app.answer_callback_query(call.id, "به سبد خرید اضافه شد")

        except Exception as e:
            print(f"❌ Error in handle_add_to_cart: {traceback.format_exc()}")

    def handle_buttons(self, call):
        """مدیریت دکمه‌های افزایش/کاهش با در نظر گرفتن واریانت"""
        try:
            data = call.data.split("_")
            action = data[0]  # increase یا decrease
            product_code = str(data[1])
            variant_id = str(data[2]) if len(data) > 2 else "0"  # 🆕 گرفتن variant_id
            
            chat_id = call.message.chat.id
            message_id = call.message.message_id

            product = Product.objects.get(code=product_code)
            profile = ProfileModel.objects.get(tel_id=chat_id)
            cart, _ = Cart.objects.get_or_create(profile=profile)

            session = SessionManager()
            user_session = session.get_user_session(chat_id, namespace="variants")
            variant_states = user_session.get(str(product_code), {})

            variants_dict = self.get_variants_dict(product.variants.all())
            selected_values = {}
            
            for i, key in enumerate(variants_dict.keys()):
                if str(i) in variant_states:
                    values_list = list(variants_dict[key])
                    selected_index = variant_states[str(i)]
                    if selected_index < len(values_list):
                        selected_values[key] = values_list[selected_index]

            # 🆕 پیدا کردن واریانت بر اساس variant_id از callback_data
            variant = None
            if variant_id != "0":
                try:
                    variant = ProductVariant.objects.get(id=variant_id, product=product)
                except ProductVariant.DoesNotExist:
                    print(f"Variant with id {variant_id} not found, using selected values")
                    # اگر variant_id پیدا نشد، از selected_values استفاده کن
                    if selected_values:
                        variant = self.get_variant_by_selected_values(product, selected_values)
            elif selected_values:
                # اگر variant_id نداریم اما selected_values داریم
                variant = self.get_variant_by_selected_values(product, selected_values)

            # 🆕 جستجوی دقیق CartItem بر اساس محصول و واریانت
            cart_item = None
            if variant:
                cart_item = CartItem.objects.filter(
                    cart=cart, 
                    product=product, 
                    variant=variant
                ).first()
            else:
                # اگر واریانت نداریم
                cart_item = CartItem.objects.filter(
                    cart=cart, 
                    product=product, 
                    variant__isnull=True
                ).first()

            should_show_initial = False

            if action == "increase":
                if not cart_item:
                    # ایجاد آیتم جدید با واریانت صحیح
                    cart_item = CartItem.objects.create(
                        cart=cart, 
                        product=product,
                        variant=variant,  # 🆕 ذخیره واریانت
                        quantity=1
                    )
                else:
                    if not variant:
                        a = {}
                        for i, (key, values) in enumerate(variants_dict.items()):
                            a[key] = values[0]
                        variant = self.get_variant_by_selected_values(product, a)
                    max_stock = variant.stock if variant else product.stock
                    if cart_item.quantity < max_stock:
                        cart_item.quantity += 1
                        cart_item.save()
                    else:
                        self.app.answer_callback_query(
                            call.id, 
                            t(call.message, "max_stock_limit", max_stock=max_stock),
                            show_alert=True
                        )
                        return
                        
            elif action == "decrease":
                if cart_item:
                    if cart_item.quantity > 1:
                        cart_item.quantity -= 1
                        cart_item.save()
                    elif cart_item.quantity == 1:
                        cart_item.quantity = 0
                        cart_item.save()
                    elif cart_item.quantity == 0:
                        cart_item.delete()
                        should_show_initial = True
                        self.app.answer_callback_query(call.id, t(call.message, "cart_item_removed"))
                else:
                    should_show_initial = True

            if should_show_initial:
                self.show_initial_state(chat_id, message_id, product)
            else:
                self.update_product_message(chat_id, message_id, product, cart)

        except ValidationError as ve:
            max_stock = variant.stock if variant else product.stock
            self.app.answer_callback_query(
                call.id, 
                t(call.message, "max_stock_limit", max_stock=max_stock), 
                show_alert=True
            )
            return
        except Exception as e:
            print(f"Error in handle_buttons: {traceback.format_exc()}")




    def update_product_message(self, chat_id, message_id, product, cart):
        """آپدیت پیام محصول با در نظر گرفتن واریانت"""
        try:
            session = SessionManager()
            user_session = session.get_user_session(chat_id, namespace="variants")
            variant_states = user_session.get(str(product.code), {})
            
            variants_dict = self.get_variants_dict(product.variants.all())
            selected_values = {}
            
            for i, key in enumerate(variants_dict.keys()):
                print(variant_states)
                if str(i) in variant_states:
                    values_list = list(variants_dict[key])
                    selected_index = variant_states[str(i)]
                    if selected_index < len(values_list):
                        selected_values[key] = values_list[selected_index]

            variant = None
            current_quantity = 0
            cart_item_exists = False
            
            if selected_values:
                variant = self.get_variant_by_selected_values(product, selected_values)

            # 🆕 جستجوی دقیق CartItem
            cart_item = None
            if variant:
                cart_item = CartItem.objects.filter(
                    cart=cart, 
                    product=product, 
                    variant=variant
                ).first()
            else:
                cart_item = CartItem.objects.filter(
                    cart=cart, 
                    product=product, 
                    variant__isnull=True
                ).first()

            if cart_item:
                current_quantity = cart_item.quantity
                cart_item_exists = True

            buttons = []
            handlers = {}

            # 🆕 اضافه کردن variant_id به callback_dataهای افزایش/کاهش
            variant_id = variant.id if variant else "0"
            buttons.extend([
                ("➕", f"increase_{self.product.code}_{variant_id}", 2),
                (f"{current_quantity}", "count", 1),
                ("➖", f"decrease_{self.product.code}_{variant_id}", 0),
            ])
            
            handlers = {
                f"increase_{product.code}_{variant_id}": self.handle_buttons,
                f"decrease_{product.code}_{variant_id}": self.handle_buttons,
            }
            
            # دکمه‌های واریانت
            for i, (key, values) in enumerate(variants_dict.items()):
                current_index = variant_states.get(str(i), 0)
                current_value = values[current_index] if current_index < len(values) else values[0]
                
                buttons.extend([
                    ("⏪", f"VarPrev_{product.code}_{i}", i * 3 + 3),
                    (f"{key}: {current_value}", f"var_{i}", i * 3 + 4),
                    ("⏩", f"VarNext_{product.code}_{i}", i * 3 + 5),
                ])
                
                handlers[f"VarPrev_{product.code}_{i}"] = self.handle_variant_navigation
                handlers[f"VarNext_{product.code}_{i}"] = self.handle_variant_navigation

            # دکمه سبد خرید
            total_cart_items = cart.total_items()
            buttons.append((f"{t("message", "menu_cart", chat_id=chat_id)} ({total_cart_items})", "view_cart", len(buttons) + 2))
            
            # لیآوت
            if variants_dict:
                button_layout = [3] + [3] * len(variants_dict) + [1]
            else:
                button_layout = [3, 1]

            stock_info = variant.stock if variant else product.stock
            if not variant:
                a = {}
                for i, (key, values) in enumerate(variants_dict.items()):
                    a[key] = values[0]
                variant = self.get_variant_by_selected_values(product, a)
            stock_info = variant.stock if variant else product.stock

            text = t("message", "order_up_to_stock", chat_id=chat_id, stock_info=stock_info)

            markup = SendMarkup(
                bot=self.app,
                chat_id=chat_id,
                text=text,
                buttons=buttons,
                button_layout=button_layout,
                handlers=handlers,
            )
            
            markup.edit(message_id)

        except Exception as e:
            print(f"❌ Error in update_product_message: {traceback.format_exc()}")



    def handle_variant_navigation(self, call):
        """مدیریت ناوبری واریانت‌ها"""
        try:
            parts = call.data.split("_")
            action_type = parts[0]
            product_code = parts[1]
            variant_index = int(parts[2])

            chat_id = call.message.chat.id
            message_id = call.message.message_id

            profile = ProfileModel.objects.get(tel_id=chat_id)
            cart, _ = Cart.objects.get_or_create(profile=profile)

            session = SessionManager()
            user_session = session.get_user_session(chat_id, namespace="variants")
            variant_states = user_session.get(str(product_code), {})

            variants_dict = self.get_variants_dict(self.product.variants.all())
            variant_keys = list(variants_dict.keys())

            if variant_index >= len(variant_keys):
                self.app.answer_callback_query(call.id, "خطا در یافتن واریانت!", show_alert=True)
                return

            current_key = variant_keys[variant_index]
            values = list(variants_dict[current_key])
            current_state = variant_states.get(str(variant_index), 0)

            if action_type == "VarPrev":
                current_state = (current_state - 1) % len(values)
            elif action_type == "VarNext":
                current_state = (current_state + 1) % len(values)

            variant_states[str(variant_index)] = current_state
            user_session[str(product_code)] = variant_states
            session.set_user_session(chat_id, user_session, namespace="variants")

            self.update_product_message(chat_id, message_id, self.product, cart)
            
            current_value = values[current_state]
            self.app.answer_callback_query(call.id, f"{current_key} به {current_value} تغییر کرد")

        except Exception as e:
            print(f"Error in handle_variant_navigation: {traceback.format_exc()}")
            self.app.answer_callback_query(call.id, "خطا در تغییر واریانت!", show_alert=True)

    def handle_comments(self, call):
        """مدیریت دکمه نظرات"""
        try:
            chat_id = call.message.chat.id
            self.app.send_message(chat_id, "صفحه نظرات محصول...")
            self.app.answer_callback_query(call.id)
        except Exception as e:
            print(f"Error in handle_comments: {traceback.format_exc()}")


############################  SEND CART  ############################    
from collections import OrderedDict
import traceback

import traceback
from telebot import types
from decimal import Decimal

# فرض می‌کنیم توابع t و کلاس‌های مورد نیاز (Cart, CartItem, ProfileModel) و SendMarkup قبلاً تعریف شده‌اند.

class SendCart(SendMarkup):
    """
    کلاس تخصصی برای نمایش و مدیریت پویای سبد خرید (Cart) در تلگرام.
    از SendMarkup برای ارسال پیام و کیبورد ارث می‌برد.
    """
    

    def __init__(self, bot, message, open_product_id=None, **kwargs):
        try:
            self.bot = bot

            # تعیین message و chat_id بر اساس نوع ورودی
            if isinstance(message, types.Message):
                self.message_obj = message
                self.chat_id = message.chat.id
                self.is_callback = False
            elif isinstance(message, types.CallbackQuery):
                self.message_obj = message.message
                self.chat_id = message.message.chat.id
                self.is_callback = True
                self.callback_query = message
            else:
                raise ValueError("Invalid message type. Must be Message or CallbackQuery")

            # دریافت پروفایل و سبد خرید
            try:
                self.profile = ProfileModel.objects.get(tel_id=self.chat_id)
                self.cart = Cart.objects.filter(profile=self.profile).first()

                if not self.cart or not self.cart.items.exists():
                    if hasattr(message, "message_id") and hasattr(message, "text"):
                        self.bot.send_message(self.chat_id, t(message, "cart_empty"))
                    else:
                        self.bot.edit_message_text(
                            chat_id=self.chat_id,
                            message_id=message.message.message_id,
                            text=t(message, "cart_empty"),
                            reply_markup=None
                        )
                    self.cart = None
                    return

                self.cart.items.filter(quantity=0).delete()

            except ProfileModel.DoesNotExist:
                self.bot.send_message(self.chat_id, "پروفایل کاربر یافت نشد!")
                self.cart = None
                return

            # مدیریت open_product_id در سشن
            session_data = session_manager.get_user_session(self.chat_id, namespace="cart") or {}

            if open_product_id is not None:
                # فقط is_open را بروزرسانی کن، بقیه سشن را حفظ کن
                session_data['is_open'] = open_product_id
                session_manager.set_user_session(self.chat_id, session_data, namespace="cart")
            else:
                if 'is_open' not in session_data:
                    session_data['is_open'] = None
                    session_manager.set_user_session(self.chat_id, session_data, namespace="cart")

            # محاسبه قیمت کل و متن
            self.total_price = sum(item.total_price() for item in self.cart.items.all())
            self.text = t(message, "cart_summary", total_price=self.total_price)

            # ایجاد دکمه‌ها و layout
            self.buttons = self._generate_buttons(self.message_obj)
            self.button_layout = self._generate_layout()

            # ثبت هندلرها
            self.handlers = self._register_handlers()

            # کش کیبورد
            self._keyboard_cache = {}

            # current_site برای لینک پرداخت
            self.current_site = 'https://intelleum.ir:8443'

            # فراخوانی سازنده والد
            super().__init__(
                bot=bot,
                chat_id=self.chat_id,
                text=self.text,
                buttons=self.buttons,
                button_layout=self.button_layout,
                handlers=self.handlers,
                **kwargs
            )

        except Exception as e:
            print(f"Error in SendCart.__init__: {e}\n{traceback.format_exc()}")
            if hasattr(self, 'chat_id'):
                self.bot.send_message(self.chat_id, t(message, "cart_load_error"))
            raise
            
    # --- متدهای کمکی برای تولید محتوا ---

    def _get_user_language(self):
        """بازیابی زبان کاربر برای ترجمه"""
        try:
            # توجه: self.chat_id در اینجا باید قبلاً توسط super().__init__ ذخیره شده باشد.
            # اگر SendMarkup این کار را نمی‌کند، باید قبل از super، self.chat_id را ذخیره کنید.
            return ProfileModel.objects.get(tel_id=self.chat_id).lang
        except Exception:
            return 'en' # زبان پیش‌فرض

    def _generate_cart_header(self, message):
        """
        1. ساخت متن اصلی پیام: 🛒 سبد خرید شما: 💰 مجموع مبلغ قابل پرداخت: فلان تومان
        """
        # این متد حالا می‌تواند از self.cart استفاده کند، زیرا در ابتدای __init__ تنظیم شده است.
        total_price = self.cart.total_price() # از متد مدل Cart استفاده می‌کند
        
        text = t(message, "cart_summary", total_price=total_price)
        
        return text


    def _get_product_items(self):
        """گروه‌بندی CartItemها بر اساس Product (نه Variant)"""
        product_items = {}
        # دسترسی به self.cart اکنون ایمن است
        for item in self.cart.items.all().select_related('product', 'variant'):
            product_id = item.product.id
            if product_id not in product_items:
                product_items[product_id] = {
                    "product": item.product,
                    "total_quantity": 0,
                    "items": []
                }
            product_items[product_id]["total_quantity"] += item.quantity
            product_items[product_id]["items"].append(item)
        return product_items

    
    def _generate_buttons(self, message):
        """
        تولید دکمه‌های محصولات و واریانت‌ها با فوکوس دقیق
        """
        buttons = {}
        product_items = self._get_product_items()
        index_counter = 1

        # session
        session_data = session_manager.get_user_session(self.chat_id, namespace="cart") or {}

        for product_id, data in product_items.items():
            product = data["product"]
            total_quantity = data["total_quantity"]

            # toggle open/close
            is_open = session_data.get("is_open") == product_id
            arrow = "▲" if is_open else "▼"

            buttons[f"{product.name} x {total_quantity} {arrow}"] = {
                'callback_data': f"cart_toggle:{product_id}",
                'index': index_counter
            }
            index_counter += 1

            if not is_open:
                continue

            # sort items
            sorted_items = sorted(data["items"], key=lambda x: x.id)

            # focused item
            focused_item_id = session_data.get(f"focused_item_{product_id}")
            focused_item = next((item for item in sorted_items if item.id == focused_item_id), None)

            if not focused_item and sorted_items:
                focused_item = sorted_items[0]
                session_data[f"focused_item_{product_id}"] = focused_item.id
                session_manager.set_user_session(self.chat_id, session_data, namespace="cart")

            if not focused_item:
                continue

            item = focused_item

            # ========================================
            #  کنترل variant و ساخت variant_info بدون خطا
            # ========================================
            variant = item.variant
            product_variants = list(product.variants.all())
            has_variants = len(product_variants) > 1

            if not has_variants:
                # محصول ساده → بدون واریانت
                variant_info = "simple_product"

            else:
                # محصول دارای واریانت است
                if variant is None:
                    # آیتم بدون واریانت → اولین واریانت انتخاب شود
                    variant = product_variants[0]

                # اکنون variant همیشه مقدار دارد
                variant_info = self._get_variant_display_info(variant)[1]

            # ========================================
            #  دکمه‌های پیمایش واریانت ONLY IF محصول واریانت دارد
            # ========================================
            if has_variants:
                buttons["<<"] = {
                    'callback_data': f"cart_prev_variant:{product_id}:{item.id}",
                    'index': index_counter,
                    'row': 'variant_nav'
                }
                index_counter += 1

                buttons[variant_info] = {
                    'callback_data': "cart_ignore",
                    'index': index_counter,
                    'row': 'variant_nav'
                }
                index_counter += 1

                buttons[">>"] = {
                    'callback_data': f"cart_next_variant:{product_id}:{item.id}",
                    'index': index_counter,
                    'row': 'variant_nav'
                }
                index_counter += 1

            # ========================================
            #  کنترل تعداد
            # ========================================
            buttons["❌"] = {
                'callback_data': f"cart_remove_product:{product_id}",
                'index': index_counter,
                'row': 'item_control'
            }
            index_counter += 1

            buttons["➖"] = {
                'callback_data': f"cart_dec:{item.id}",
                'index': index_counter,
                'row': 'item_control'
            }
            index_counter += 1

            buttons[f"{item.quantity}"] = {
                'callback_data': "cart_ignore",
                'index': index_counter,
                'row': 'item_control'
            }
            index_counter += 1

            buttons["➕"] = {
                'callback_data': f"cart_inc:{item.id}",
                'index': index_counter,
                'row': 'item_control'
            }
            index_counter += 1

        # checkout button
        buttons[t(message, "checkout")] = {
            'callback_data': "cart_checkout",
            'index': index_counter
        }

        return buttons




    def _handle_variant_nav(self, call, product_id, current_item_id, direction):
        """
        پیمایش بین واریانت‌های موجود در سبد خرید
        """
        try:
            variants_in_cart = self._get_product_variants_in_cart(product_id)
            if not variants_in_cart:
                self.bot.answer_callback_query(call.id, "هیچ واریانتی برای این محصول موجود نیست", show_alert=True)
                return

            # پیدا کردن ایندکس واریانت فعلی
            current_index = next((i for i, v in enumerate(variants_in_cart) if v['item_id'] == int(current_item_id)), 0)

            # محاسبه ایندکس جدید
            if direction == "next":
                new_index = (current_index + 1) % len(variants_in_cart)
            else:
                new_index = (current_index - 1) % len(variants_in_cart)

            new_variant_data = variants_in_cart[new_index]

            # بروزرسانی سشن کامل
            session_data = session_manager.get_user_session(self.chat_id, namespace="cart") or {}
            session_data[f"focused_variant_{product_id}"] = new_variant_data['variant_id']
            session_data[f"focused_item_{product_id}"] = new_variant_data['item_id']
            session_manager.set_user_session(self.chat_id, session_data, namespace="cart")

            # بروزرسانی منو با سشن کامل
            new_cart_menu = SendCart(
                bot=self.bot,
                message=call.message,
                open_product_id=product_id  # محصول باز می‌ماند
            )
            new_cart_menu.edit(call.message.message_id)

            self.bot.answer_callback_query(call.id, f"تغییر به: {new_variant_data['variant_info']}")

        except Exception as e:
            print(f"Error in _handle_variant_nav: {e}\n{traceback.format_exc()}")
            self.bot.answer_callback_query(call.id, "خطا در تغییر واریانت", show_alert=True)




    def _generate_layout(self):
        """
        تولید طرح‌بندی (Layout) برای کیبورد اینلاین
        """
        layout = []
        
        # ردیف‌های کالاها (هر ردیف یک دکمه)
        for button_text, data in self.buttons.items():
            if data['callback_data'].startswith("cart_toggle"):
                layout.append(1) # دکمه اصلی کالا
            elif 'row' in data:
                # ردیف‌های پیمایش و کنترل واریانت
                if data['row'] == 'variant_nav' and 3 not in layout:
                    layout.append(3) # << variant >>
                elif data['row'] == 'item_control' and 4 not in layout:
                    layout.append(4) # X - Q +
                elif data['row'] == 'control' and 4 not in layout:
                    layout.append(4) # X - Q +
            
        # دکمه نهایی
        layout.append(1) # تکمیل خرید
        
        return layout


    # --- متدهای مدیریت کلیک (Handlers) ---

    def _register_handlers(self):
        """ثبت هندلرهای مربوط به Callback Dataها"""
        handlers = {
            # "cart_prev_variant": self._handle_prev_variant,
            # "cart_next_variant": self._handle_next_variant,
            "cart_remove_product": self._handle_remove_product,
            "cart_dec": self._handle_inc_dec,
            "cart_checkout": self._handle_checkout,
            "cart_ignore": lambda call, *args: None,
        }
        
        # هندلر برای باز و بسته کردن (cart_toggle:product_id)
        # و همچنین برای افزایش/کاهش و پیمایش، که با متد کلی مدیریت می‌شوند
        
        return handlers




    def _handle_checkout(self, call):
        """هندل کردن دکمه تکمیل خرید"""
        self.invoice(call)
        # منطق انتقال به صفحه پرداخت

    def _handle_ignore(self, call):
        """هندل کردن دکمه‌های صرفاً نمایشی"""
        self.bot.answer_callback_query(call.id, "هندل کردن دکمه‌های صرفاً نمایشی")
    
    def _handle_toggle(self, call, product_id):
        """
        3. باز و بسته کردن منوی واریانت یک محصول
        """
        # اگر در حال حاضر باز است، آن را ببند (تنظیم روی None)
        try:   
            data = session_manager.get_user_session(self.chat_id, namespace="cart")
            if data.get("is_open") == product_id:
                new_open_product_id = None

            else:
                # اگر بسته است یا محصول دیگری باز است، این محصول را باز کن
                new_open_product_id = product_id
    
            # ساخت یک آبجکت جدید با وضعیت باز جدید
            session_data = session_manager.get_user_session(self.chat_id, namespace="cart")
            session_data["is_open"] = new_open_product_id 
            session_manager.set_user_session(self.chat_id, session_data, namespace="cart")
            new_cart_menu = SendCart(self.bot, call.message, new_open_product_id)
            new_cart_menu.edit(call.message.message_id)
            self.bot.answer_callback_query(call.id)
        except:
            print(traceback.format_exc())
        
    def _handle_remove_product(self, call, product_id):
        """
        4. حذف کل کالا (تمام آیتم‌های آن) از سبد خرید
        """
        try:
            # حذف تمام CartItemهای مربوط به این Product در این Cart
            removed_count = self.cart.items.filter(product_id=product_id).delete()
            if removed_count[0] > 0:
                self.bot.answer_callback_query(call.id, f"{removed_count[0]} آیتم حذف شد.")
                # رفرش سبد خرید (حالت باز را به None تغییر می‌دهد)
                session_data = session_manager.get_user_session(self.chat_id, namespace="cart")
                if session_data.get("is_open") == product_id:
                    session_data["is_open"] = None
                    session_manager.set_user_session(self.chat_id, session_data, namespace="cart")
                if not self.cart.items.filter(quantity__gt=0).exists():
                    session_manager.reset_user_session(self.chat_id, namespace="cart")
                    # سبد خرید خالی است، پیام موجود را ویرایش کن
                    self.bot.edit_message_text(
                        chat_id=self.chat_id,
                        message_id=call.message.message_id,
                        text=t(call.message, "cart_empty"),
                        reply_markup=None
                    )
                    return
                new_cart_menu = SendCart(self.bot, call.message, open_product_id=None)
                new_cart_menu.edit(call.message.message_id)
            else:
                self.bot.answer_callback_query(call.id, "این کالا قبلاً حذف شده است.", show_alert=True)

        except Exception as e:
            print(f"Error removing product from cart: {e}")
            self.bot.answer_callback_query(call.id, "خطا در حذف کالا.", show_alert=True)
            

    def _handle_inc_dec(self, call, action, item_id):
        try:
            item = CartItem.objects.get(id=item_id, cart=self.cart)
            old_quantity = item.quantity

            # افزایش یا کاهش
            if action == 'inc':
                max_stock = item.variant.stock if item.variant else item.product.stock
                if old_quantity < max_stock:
                    item.quantity += 1
                    item.save()
                    self.bot.answer_callback_query(call.id)
                else:
                    self.bot.answer_callback_query(call.id, f"⚠️ بیش از {max_stock} عدد موجود نیست!", show_alert=True)
                    return

            elif action == 'dec':
                new_quantity = old_quantity - 1
                if new_quantity <= 0:
                    item.delete()
                    self.bot.answer_callback_query(call.id, t(call.message, "cart_item_removed"))
                else:
                    item.quantity = new_quantity
                    item.save()
                    self.bot.answer_callback_query(call.id)
                    
            # بررسی خالی بودن سبد خرید
            if not self.cart.items.filter(quantity__gt=0).exists():
                session_manager.reset_user_session(self.chat_id, namespace="cart")
                # سبد خرید خالی است، پیام موجود را ویرایش کن
                self.bot.edit_message_text(
                    chat_id=self.chat_id,
                    message_id=call.message.message_id,
                    text=t(call.message, "cart_empty"),
                    reply_markup=None
                )
                return  # نیازی به SendCart جدید نیست

            # اگر هنوز آیتمی وجود دارد، رفرش منو با حفظ open_product_id
            new_cart_menu = SendCart(self.bot, call.message, item.product.id)
            new_cart_menu.edit(call.message.message_id)

        except CartItem.DoesNotExist:
            self.bot.answer_callback_query(call.id, "آیتم پیدا نشد.", show_alert=True)
        except ValidationError as e:
            error_msg = list(e.message_dict.values())[0][0] if e.message_dict else str(e)
            self.bot.answer_callback_query(call.id, error_msg, show_alert=True)
        except Exception as e:
            print(f"Error handling cart item quantity: {traceback.format_exc()}")
            self.bot.answer_callback_query(call.id, "خطای نامشخص.", show_alert=True)

    
    



    def _get_variant_display_info(self, variant):
        """دریافت اطلاعات نمایشی واریانت"""
        if not variant:
            return "بدون واریانت"
        
        value = []
        for option_value in variant.values.all().order_by('option__name'):
            value.append(f"{option_value.option.name}: {option_value.value}")

        item = [ov.value for ov in variant.values.all().order_by('option__name')]
        
        values = "، ".join(value)
        items = "/ ".join(item)
        return values, items


    def _get_product_variants_in_cart(self, product_id):
        """دریافت واریانت‌های یک محصول که در سبد خرید موجود هستند"""
        # if product_id in self._cached_product_variants:
        #     return self._cached_product_variants[product_id]
        
        variants_in_cart = []
        for item in self.cart.items.filter(product_id=product_id, quantity__gt=0):
            if item.variant:
                variants_in_cart.append({
                    'item_id': item.id,
                    'variant_id': item.variant.id,
                    'variant_info': self._get_variant_display_info(item.variant)[0],
                    'quantity': item.quantity
                })
        
        # self._cached_product_variants[product_id] = variants_in_cart
        return variants_in_cart



    def handle_callback(self, call):
        """مدیریت کلیک‌های پویا"""
        callback_data = call.data
        
        # مدیریت دکمه‌های اصلی واریانت (توگل، افزایش/کاهش، حذف، پیمایش)
        parts = callback_data.split(":")
        action = parts[0]
        
        if action == "cart_toggle" and len(parts) == 2:
            self._handle_toggle(call, int(parts[1]))
        elif action == "cart_remove_product" and len(parts) == 2:
            self._handle_remove_product(call, int(parts[1]))
        elif action == "cart_inc" and len(parts) == 2:
            self._handle_inc_dec(call, 'inc', int(parts[1]))
        elif action == "cart_dec" and len(parts) == 2:
            self._handle_inc_dec(call, 'dec', int(parts[1]))
        elif action == "cart_next_variant" and len(parts) == 3:
            self._handle_variant_nav(call, int(parts[1]), int(parts[2]), 'next')
        elif action == "cart_prev_variant" and len(parts) == 3:
            self._handle_variant_nav(call, int(parts[1]), int(parts[2]), 'prev')
        else:
            # مدیریت سایر هندلرهای ثبت شده (مانند checkout یا ignore)
            super().handle_callback(call)

    def invoice(self, update):
        try:
            # تشخیص نوع ورودی
            if isinstance(update, types.CallbackQuery):
                chat_id = update.message.chat.id
                message = update.message
                message_id = update.message.message_id
                is_callback = True
            elif isinstance(update, types.Message):
                chat_id = update.chat.id
                message = update
                message_id = update.message_id
                is_callback = False
            else:
                print("❌ update type not supported in invoice()")
                return

            profile = ProfileModel.objects.get(tel_id=chat_id)
            cart = Cart.objects.get(profile=profile)
            cart_items = CartItem.objects.filter(cart=cart)

            total_price = sum(item.total_price() for item in cart_items)

            invoice_text = t(message, "order_invoice")
            for index, item in enumerate(cart_items, start=1):
                invoice_text += f"{index}) {item.product.name}  -  "
                invoice_text += f"{item.product.final_price:,.0f} x {item.quantity}\n\n"
            invoice_text += t(message, "total_amount", total_price=total_price)

            address = Address.objects.filter(profile=profile, shipping_is_active=True).first()
            try:
                line1 = address.shipping_line1[:10] + '...' if len(address.shipping_line1)>10 else address.shipping_line1
            except Exception as e:
                line1 = ''
            address_text = (f"{line1}, {address.shipping_city_name}, {address.shipping_province_name}, {address.shipping_country_name}"
                            if address else ' --- ')
            if len(address_text)>40:
                address_text = address_text[:40] + "..."
                
            phone_text = (f"{profile.phone}" if profile.phone else ' --- ')

            payment_link = self.pay(update)

            # اصلاح شده: اطمینان از اینکه callback_data هیچگاه خالی نباشد
            buttons = {
                t(message, "address", address_text=address_text): {"callback_data": "address", "index": 1},
                t(message, "contact_number", phone_text=phone_text): {"callback_data": "phone", "index": 2},
            }

            # اضافه کردن دکمه پرداخت با فرمت صحیح
            if address and profile.phone:
                buttons[t(message, "pay_and_purchase")] = {"url": payment_link, "index": 3}
            else:
                buttons[t(message, "pay_and_purchase")] = {"callback_data": "phone_address_required", "index": 3}

            self.markup = SendMarkup(
                bot=self.bot,
                chat_id=chat_id,
                text=invoice_text,
                buttons=buttons,
                button_layout=[1, 1, 1],
                handlers={
                    "address": lambda call: SendLocation(self.bot, call.message).show_addresses(),
                    "phone_address_required": lambda call: self.bot.answer_callback_query(call.id, "⚠️ لطفاً ابتدا آدرس و شماره تماس را تکمیل کنید.")
                }
            ) 
            if isinstance(update, types.CallbackQuery):
                self.markup.edit(message_id)
            elif isinstance(update, types.Message):
                self.markup.send()

            if is_callback:
                self.bot.answer_callback_query(update.id, t(message, "processing_payment"))

        except Exception as e:
            print(f"Error in invoice: {e}\n{traceback.format_exc()}")
            self.bot.send_message(chat_id, "❌ خطایی در نمایش فاکتور رخ داد.")


    def pay(self, update):

        # تشخیص نوع ورودی
        if isinstance(update, types.CallbackQuery):
            chat_id = update.message.chat.id
            message_id = update.message.message_id
            is_callback = True
        elif isinstance(update, types.Message):
            chat_id = update.chat.id
            message_id = update.message_id
            is_callback = False
        else:
            print("❌ update type not supported in invoice()")
            return


        # 2. ایجاد شناسه یکتا برای پرداخت
        payment_id = str(uuid.uuid4())

        # 3. ذخیره داده در کش
        cache.set(
            f'payment_{payment_id}',
            {'tel_id': chat_id},
            timeout=settings.PAYMENT_LINK_TIMEOUT
        )

        # 4. ساخت لینک پرداخت
        payment_link = f"{settings_current_site}/buy?pid={payment_id}"

        return payment_link



############################  SEND LOCATION  ############################

class SendLocation:
    def __init__(self, app, message_or_call):
        """
        مقداردهی اولیه کلاس
        :param app: شیء بات
        :param message_or_call: می‌تواند Message یا CallbackQuery باشد
        """
        try:
            self.app = app
            self.session_manager = session_manager
            self.user_id = (message_or_call.from_user.id if hasattr(message_or_call, 'from_user') else message_or_call.message.from_user.id)
            self.chat_id = message_or_call.chat.id if hasattr(message_or_call, 'chat') else message_or_call.message.chat.id
            self.message = message_or_call if isinstance(message_or_call, types.Message) else message_or_call.message
            self.profile = ProfileModel.objects.get(tel_id=self.chat_id)
            self.user_addresses = Address.objects.filter(profile=self.profile)
            self.active_address = Address.objects.filter(profile=self.profile, shipping_is_active=True).first()
        except Exception as e:
            error_details = traceback.format_exc()
            custom_message = f"Error in SendLocation init: {e}\nDetails:\n{error_details}"
            print(custom_message)
            self.app.send_message(self.chat_id, t("message", "error_address_info", chat_id=self.chat_id))

    def show_addresses(self, call=None):
        """
        نمایش لیست آدرس‌های کاربر
        :param call: در صورتی که از طریق callback فراخوانی شده باشد
        """
        try:
            # متن پیام
            text = t(self.message, "your_addresses")

            # ساخت دکمه‌های آدرس‌ها
            buttons = {}

            for i, address in enumerate(self.user_addresses, start=1):
                line1 = address.shipping_line1[:15] + '...' if len(address.shipping_line1)>15 else address.shipping_line1
                btn_text = f"{i}. {line1}, {address.shipping_city_name}, {address.shipping_province_name}, {address.shipping_country_name}"
                if len(btn_text)>40:
                    btn_text = btn_text[:40] + "..."
                if address == self.active_address:
                    btn_text += " ⭐️"  # نشانگر آدرس فعال
                buttons[btn_text] = (f"show_address_{address.id}", i)

            # دکمه‌های پایه
            buttons[t(self.message, "add_new_address")] = ("add_new_address", len(buttons)+1)
            buttons[t(self.message, "close")] = ("close_addresses", len(buttons)+2)

            handlers = {
                "add_address": self.handle_add_address,
                "close_address": self.handle_close,
            }

            # اضافه کردن هندلرهای آدرس‌ها
            for address in self.user_addresses:
                handlers[f"address_{address.id}"] = lambda addr, c=address: self.show_single_address(addr, c, chat_id=call.message.chat.id)

            # ایجاد کیبورد
            markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=buttons,
                button_layout=[1]*len(self.user_addresses) + [2],
                handlers=handlers
            )

            # ارسال یا ویرایش پیام
            if call:
                markup.edit(call.message.message_id)  # ویرایش پیام موجود
            else:
                markup.send()  # ارسال پیام جدید

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "address_display_error"))

    def show_single_address(self, address, call=None, chat_id=None):
        """
        نمایش جزئیات یک آدرس خاص
        :param call: شیء callback
        :param address: آدرس انتخابی
        """
        try:
            # متن پیام
            text = f"{t("message", "your_addresses", chat_id=chat_id)}{address.shipping_line1}\n\n"
            text += f"{t("message", "city_label", chat_id=chat_id)} {address.shipping_city_name}\n"
            text += f"{t("message", "province_label", chat_id=chat_id)} {address.shipping_province_name}\n"
            text += f"{t("message", "country_label", chat_id=chat_id)} {address.shipping_country_name}\n"
            text += f"{t("message", "postal_code_label", chat_id=chat_id)} {address.shipping_zip_code or t("message", "not_registered")}"

            # دکمه‌های مدیریت
            buttons = {
                t("message", 'change_location', chat_id=chat_id): (f'change_location_{address.id}', 1),
                t("message", "send_to_this_address", chat_id=chat_id): (f"select_address_{address.id}", 2),
                t("message", "edit_address", chat_id=chat_id): (f"change_address_{address.id}", 3),
                t("message", "edit_postal_code", chat_id=chat_id): (f"change_postal_{address.id}", 4),
                t("message", "back_button", chat_id=chat_id): ("back_to_addresses", 5),
                t("message", "delete_address", chat_id=chat_id): (f"delete_address_{address.id}", 6)
            }

            markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=buttons,
                button_layout=[1, 1, 1, 1, 2],
                handlers={
                    f"change_location_{address.id}": lambda c: self.change_location(c, address),
                    f"select_address_{address.id}": lambda c: self.select_address(c, address),
                    f"change_address_{address.id}": lambda c: self.change_address_text(c, address),
                    f"change_postal_{address.id}": lambda c: self.change_postal(c, address),
                    "back_to_addresses": lambda c: self.show_addresses(c),
                    f"delete_address_{address.id}": lambda c: self.delete_address(c, address)
                }
            )

            # ارسال یا ویرایش پیام
            if call:
                markup.edit(call.message.message_id)  # ویرایش پیام موجود
            else:
                markup.send()  # ارسال پیام جدید

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_single_address: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t("message", "address_display_error", chat_id=chat_id))

    # --- متدهای مدیریت عملیات ---

    def handle_add_address(self, call):
        """افزودن آدرس جدید"""
        try:
            self.app.send_message(call.message.chat.id, t(call.message, "please_send_new_address"))
            # اینجا می‌توانید از register_next_step_handler استفاده کنید
        except Exception as e:
            print(f"Error in handle_add_address: {traceback.format_exc()}")
            self.app.send_message(call.message.chat.id, t(call.message, "error_add_address"))

    def handle_close(self, call):
        """بستن پنجره آدرس‌ها"""
        try:
            self.app.delete_message(call.message.chat.id, call.message.message_id)
            self.session_manager.reset_user_session(call.message.chat.id, namespace="address")
        except Exception as e:
            print(f"Error in handle_close: {traceback.format_exc()}")

    def change_location(self, call, address):
        """تغییر موقعیت مکانی"""
        try:
            self.app.send_message(call.message.chat.id,
                                  t(call.message, "change_location_prompt"),
                                  reply_markup=types.ReplyKeyboardMarkup(
                                      resize_keyboard=True
                                  ).add(types.KeyboardButton(t(call.message, "share_location"), request_location=True)))
            # ذخیره آدرس برای مرحله بعد
            # اینجا می‌توانید از register_next_step_handler استفاده کنید
        except Exception as e:
            print(f"Error in change_location: {traceback.format_exc()}")
            self.app.send_message(call.message.chat.id, t(call.message, "change_location_error"))


    def delete_address(self, call, address):
        """حذف آدرس"""
        try:
            address.delete()
            self.app.answer_callback_query(call.id, t(call.message, "address_deleted"))
            self.show_addresses(call)
        except Exception as e:
            print(f"Error in delete_address: {traceback.format_exc()}")
            self.app.answer_callback_query(call.id, t(call.message, "address_delete_error"))

    def add_new_address(self, call):
        try:

            text = t(call.message, "choose_address_input_method")


            buttons = {
                t(call.message, "manual_entry"): (f"manual_add_address", 1),
                t(call.message, "send_location"): (f"send_location_add_address", 2),
            }

            handlers = {
                "manual_add_address": self.manual_add_address,
                "send_location_add_address": self.send_location_add_address,
            }

            data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")
            data["state"] = "add_new_address"
            self.session_manager.set_user_session(call.message.chat.id, data, namespace="address")



            # ایجاد کیبورد
            markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=buttons,
                button_layout=[2],
                handlers=handlers
            )

            markup.edit(call.message.message_id)

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in add_new_address: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "address_display_error"))

    def manual_add_address(self, call):
        try:
            text = t(call.message, "select_country")
            items = [item for item in get_country_choices(self.profile.lang)]
            items_name = [item[1] for item in get_country_choices(self.profile.lang)]
            items_code = [item[0] for item in get_country_choices(self.profile.lang)]
            paginator = InlineKeyboardPaginator(user_id=self.user_id, items=items_name, per_page=24, row_size=3, remember_last_page=True)
            buttons, layout = paginator.get_buttons_for_sendmarkup()

            handlers = {"prev": self.handle_prev, "next": self.handle_next}

            for code, country in items:
                if country in buttons:
                    buttons[country]["callback_data"] = f"country_{code}"
                    handlers[f'country_{code}'] = self.handle_picked_country

            buttons['🔙'] = {'callback_data': '_back', 'index': len(buttons)+1}
            handlers["_back"] = self.handle_previous

            buttons[t(call.message, "close")] = {'callback_data': 'address_close', 'index': len(buttons)+2}
            handlers["address_close"] = self.handle_close
            layout.append(2)


            data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")
            data["state"] = "address_selection_country"
            self.session_manager.set_user_session(call.message.chat.id, data, namespace="address")



            # ایجاد کیبورد
            markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=buttons,
                button_layout=layout,
                handlers=handlers
            )


            markup.edit(call.message.message_id)


        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "address_display_error"))


    def handle_prev(self, call):
        try:
            session = self.session_manager.get_user_session(call.message.chat.id, namespace="address")

            paginator = InlineKeyboardPaginator.load_from_redis(self.user_id)
            paginator.prev_page()
            handlers = {"prev": self.handle_prev, "next": self.handle_next}
            if session['state'] == "address_selection_country":
                text = t(call.message, "select_country")
                items = [item for item in get_country_choices(self.profile.lang)]
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                for code, country in items:
                    if country in buttons:
                        buttons[country]["callback_data"] = f"country_{code}"
                        handlers[f'country_{code}'] = self.handle_picked_country

            elif session['state'] == "address_selection_province":
                text = t(call.message, "select_province")
                items = [item for item in get_province_choices(session["selected_country"], self.profile.lang)]
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                for code, province in items:
                    if province in buttons:
                        buttons[province]["callback_data"] = f"province_{code}"
                        handlers[f'province_{code}'] = self.handle_picked_province

            elif session['state'] == "address_selection_city":
                text = t(call.message, "select_city")
                items = [item for item in get_city_choices(session["selected_country"], session["selected_province"], self.profile.lang)]
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                for code, city in items:
                    if city in buttons:
                        buttons[city]["callback_data"] = f"city_{code}"
                        handlers[f'city_{code}'] = self.handle_picked_city
            
            buttons['🔙'] = {'callback_data': '_back', 'index': len(buttons)+1}
            handlers["_back"] = self.handle_previous

            buttons[t(call.message, "close")] = {'callback_data': 'address_close', 'index': len(buttons)+2}
            handlers["address_close"] = self.handle_close
            layout.append(2)


            SendMarkup(bot=self.app, chat_id=self.chat_id, text=text, buttons=buttons, button_layout=layout, handlers=handlers).edit(
                call.message.message_id)

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "address_display_error"))

    def handle_next(self, call):
        try:
            data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")
            paginator = InlineKeyboardPaginator.load_from_redis(self.user_id)
            paginator.next_page()
            handlers = {"prev": self.handle_prev, "next": self.handle_next}
            if data['state'] == "address_selection_country":
                text = t(call.message, "select_country")
                items = [item for item in get_country_choices(self.profile.lang)]
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                for code, country in items:
                    if country in buttons:
                        buttons[country]["callback_data"] = f"country_{code}"
                        handlers[f'country_{code}'] = self.handle_picked_country

            elif data['state'] == "address_selection_province":
                text = t(call.message, "select_province")
                items = [item for item in get_province_choices(data["selected_country"], self.profile.lang)]
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                for code, province in items:
                    if province in buttons:
                        buttons[province]["callback_data"] = f"province_{code}"
                        handlers[f'province_{code}'] = self.handle_picked_province

            elif data['state'] == "address_selection_city":
                text = t(call.message, "select_city")
                items = [item for item in get_city_choices(data["selected_country"], data["selected_province"], self.profile.lang)]
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                for code, city in items:
                    if city in buttons:
                        buttons[city]["callback_data"] = f"city_{code}"
                        handlers[f'city_{code}'] = self.handle_picked_city
            
            buttons['🔙'] = {'callback_data': '_back', 'index': len(buttons)+1}
            handlers["_back"] = self.handle_previous

            buttons[t(call.message, "close")] = {'callback_data': 'address_close', 'index': len(buttons)+2}
            handlers["address_close"] = self.handle_close
            layout.append(2)


            SendMarkup(bot=self.app, chat_id=call.message.chat.id, text=text, buttons=buttons, button_layout=layout, handlers=handlers).edit(call.message.message_id)
            data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "address_display_error"))

    def handle_picked_country(self, call):
        import math
        per_page = 15
        row_size = 3
        if call.data == '_back':
            session = self.session_manager.get_user_session(call.message.chat.id, namespace="address")
            country = session['selected_country']
        else:
            country = call.data.split("_")[-1]
        data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")
        data["state"] = "address_selection_province"
        data["selected_country"] = f"{country}"
        self.session_manager.set_user_session(call.message.chat.id, data, namespace="address")

        try:

            items = [item for item in get_province_choices(country, self.profile.lang)]
            items_name = [item[1] for item in get_province_choices(country, self.profile.lang)]
            items_code = [item[0] for item in get_province_choices(country, self.profile.lang)]
            handlers = {"prev": self.handle_prev, "next": self.handle_next}
            buttons = {}
            text = t(call.message, "select_province")

            if len(items)>per_page:
                paginator = InlineKeyboardPaginator(user_id=self.user_id, items=items_name, per_page=per_page, row_size=row_size, remember_last_page=True)
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                
                for code, province in items:
                    if province in buttons:
                        buttons[province]["callback_data"] = f"province_{code}"
                    handlers[f'province_{code}'] = self.handle_picked_province


            else:
                counter = 0
                for code, province in items:
                    counter +=1
                    buttons[f"{province}"] = {"callback_data": f"province_{code}", "index": counter}
                    handlers[f'province_{code}'] = lambda c: self.handle_picked_province(c)
                layout = [row_size for i in range(math.floor(len(items)/row_size))]
                if len(items)%row_size:
                    for i in range(len(items)%row_size):
                        layout.append(len(items)%row_size)


            buttons['🔙'] = {'callback_data': '_back', 'index': len(buttons)+1}
            handlers["_back"] = self.handle_previous

            buttons[t(call.message, "close")] = {'callback_data': 'address_close', 'index': len(buttons)+2}
            handlers["address_close"] = self.handle_close
            layout.append(2)


            markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=buttons,
                button_layout=layout,
                handlers=handlers
            )


            markup.edit(call.message.message_id)

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "address_display_error"))
  


    def handle_picked_province(self, call):
        per_page = 15
        row_size = 3
        province = call.data.split("_")[-1]
        data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")

        data["state"] = "address_selection_city"
        data["selected_province"] = f"{province}"

        self.session_manager.set_user_session(call.message.chat.id, data, namespace="address")


        try:

            items = [item for item in get_city_choices(data["selected_country"], province, self.profile.lang)]
            items_name = [item[1] for item in get_city_choices(data["selected_country"], province, self.profile.lang)]
            items_code = [item[0] for item in get_city_choices(data["selected_country"], province, self.profile.lang)]
            handlers = {"prev": self.handle_prev, "next": self.handle_next}
            buttons = {}
            text = t(call.message, "select_city")

            if len(items)>per_page:
                paginator = InlineKeyboardPaginator(user_id=self.user_id, items=items_name, per_page=per_page, row_size=row_size, remember_last_page=True)
                buttons, layout = paginator.get_buttons_for_sendmarkup()
                
                for code, city in items:
                    if city in buttons:
                        buttons[city]["callback_data"] = f"city_{code}"
                    handlers[f'city_{code}'] = self.handle_picked_city


            else:
                counter = 0
                for code, city in items:
                    counter +=1
                    buttons[f"{city}"] = {"callback_data": f"city_{code}", "index": counter}
                    handlers[f'city_{code}'] = self.handle_picked_city
                layout = [row_size for i in range(math.floor(len(items)/row_size))]
                if len(items)%row_size:
                    for i in range(len(items)%row_size):
                        layout.append(len(items)%row_size)

            buttons['🔙'] = {'callback_data': '_back', 'index': len(buttons)+1}
            handlers["_back"] = self.handle_previous

            buttons[t(call.message, "close")] = {'callback_data': 'address_close', 'index': len(buttons)+2}
            handlers["address_close"] = self.handle_close
            layout.append(2)


            markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=buttons,
                button_layout=layout,
                handlers=handlers
            )



            markup.edit(call.message.message_id)
            data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "address_display_error"))


    def handle_picked_city(self, call):

        data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")

        city = call.data.split("_")[-1]

        data["state"] = "address_selection_street"
        data["selected_city"] = f"{city}"
        data["old_message"] = call.message.message_id

        self.session_manager.set_user_session(call.message.chat.id, data, namespace="address")


        text = t(call.message, "enter_address_details")

        markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=None,
                button_layout=None,
                handlers=None
            )


        markup.edit(call.message.message_id) 
        data = self.session_manager.get_user_session(call.message.chat.id, namespace="address")


        

    def handle_picked_street(self, message):
        data = self.session_manager.get_user_session(message.chat.id, namespace="address")

        data["state"] = "address_selection_zipcode"
        data["selected_address_line1"] = f"{message.text}"

        self.session_manager.set_user_session(message.chat.id, data, namespace="address")

        text = t(message, "enter_zip_code")
        
        self.app.delete_message(message.chat.id, message.message_id)
        self.app.delete_message(message.chat.id, data["old_message"])

        message = self.app.send_message(self.chat_id, text)

        data["old_message"] = message.message_id

        self.session_manager.set_user_session(message.chat.id, data, namespace="address")


    def handle_picked_zipcode(self, message):
        try:
            data = self.session_manager.get_user_session(message.chat.id, namespace="address")

            # change postal of address previously created
            if (type(data.get("change_postal")) == list) and (data.get("change_postal")[0]):

                address = Address.objects.get(id=data.get("change_postal")[1])
                address.shipping_zip_code = int(message.text)
                address.save()

            else:
                data["state"] = "address_selection_zipcode"
                data["selected_zipcode"] = f"{message.text}"

                self.session_manager.set_user_session(message.chat.id, data, namespace="address")
            
                data = self.session_manager.get_user_session(message.chat.id, namespace="address")
                if "change_address" in list(data.keys()) and data['change_address'][0]:
                    address = Address.objects.get(id=data['change_address'][1])
                    address.shipping_line1 = data["selected_address_line1"]
                    address.shipping_country=data["selected_country"]
                    address.shipping_province=data["selected_province"]
                    address.shipping_city=data["selected_city"]
                    address.shipping_zip_code=data["selected_zipcode"]
                    address.save()
                    self.app.delete_message(message.chat.id, message.message_id)
                    self.app.delete_message(message.chat.id, data["old_message"])
                else:
                    address = Address.objects.create(profile=self.profile, shipping_line1=data["selected_address_line1"], shipping_country=data["selected_country"], shipping_province=data["selected_province"], shipping_city=data["selected_city"], shipping_zip_code=data["selected_zipcode"], shipping_is_active=True) 
                    address.save()
                    self.app.delete_message(message.chat.id, message.message_id)
                    self.app.delete_message(message.chat.id, data["old_message"])

            if (type(data.get("change_postal")) == list) and (data.get("change_postal")[0]):
                self.show_single_address(address=address, chat_id=message.chat.id)
                data['change_postal'] = None
                self.session_manager.set_user_session(message.chat.id, data, namespace="address")   
                return
            if not data.get('from my postal address'):
                SendCart(self.app, self.message).invoice(self.message)
            else:
                self.show_addresses()
            self.session_manager.reset_user_session(message.chat.id, namespace="address")

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(message, "address_display_error"))


    def select_address(self, call, address):
        try:
            self.app.answer_callback_query(call.id, t(call.message, "address_set_active"), show_alert=True)
            address.shipping_is_active = True
            address.save()
        except Exception as e:
            print(e)


    def change_postal(self, call):
        try:
            
            pass
        except Exception as e:
            print(f"Error in change_postal: {traceback.format_exc()}")


    def handle_previous(self, call):
        pass


    def send_location_add_address(self):
        pass


##############################################    SEND PHONE    ##############################################


class SendPhone:
    def __init__(self, app, message_or_call):
        """
        مقداردهی اولیه کلاس
        :param app: شیء بات
        :param message_or_call: می‌تواند Message یا CallbackQuery باشد
        """
        try:
            self.app = app
            self.session_manager = session_manager
            self.user_id = (message_or_call.from_user.id if hasattr(message_or_call, 'from_user') else message_or_call.message.from_user.id)
            self.chat_id = message_or_call.chat.id if hasattr(message_or_call, 'chat') else message_or_call.message.chat.id
            self.message = message_or_call if isinstance(message_or_call, types.Message) else message_or_call.message
            self.profile = ProfileModel.objects.get(tel_id=self.chat_id)
            self.user_phone = self.profile.phone
        except Exception as e:
            error_details = traceback.format_exc()
            custom_message = f"Error in SendPhone init: {e}\nDetails:\n{error_details}"
            print(custom_message)
            self.app.send_message(self.chat_id, t(message_or_call.message, "contact_info_error"))

    def take_phone(self, call):
        try:
            text = t(call.message, "enter_phone_number")
            text += t(call.message, "phone_example")

            markup = SendMarkup(
                bot=self.app,
                chat_id=self.chat_id,
                text=text,
                buttons=None,
                button_layout=None,
                handlers=None
            )

            data = {"state": "take_phone", "old_message": call.message.message_id}
            self.session_manager.set_user_session(call.message.chat.id, data, namespace="phone")



            # ارسال یا ویرایش پیام
            if call:
                markup.edit(call.message.message_id)  # ویرایش پیام موجود
            else:
                markup.send()  # ارسال پیام جدید


        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(call.message, "contact_info_error"))


    def really_take_phone(self, message):
        try:
            profile = ProfileModel.objects.get(tel_id=message.chat.id)

            try:
                user_cart = Cart.objects.get(profile=profile)
            except Cart.DoesNotExist:
                app.send_message(message.chat.id, t(message, "cart_is_empty"))
                return
            data = self.session_manager.get_user_session(message.chat.id, namespace="phone")
            self.app.delete_message(self.chat_id, message.message_id)
            self.app.delete_message(self.chat_id, data["old_message"])

            num = int(message.text)
            self.profile.phone = num
            self.profile.save()
            SendCart(self.app, self.message).invoice(self.message)
            self.session_manager.reset_user_session(message.chat.id, namespace="phone")


        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in show_addresses: {e}\n{error_details}")
            self.app.send_message(self.chat_id, t(message, "contact_info_error"))


##############################################    PRODUCT LIST    ##############################################

import base64
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO
import traceback
from django.db.models import Prefetch, Count, Sum, Avg
import redis
import json
from datetime import datetime, timedelta
import threading
from products.models import Product, ProductVariant, Category, Store
from accounts.models import ProfileModel

class AdvancedProductExporter:
    """کلاس کامل برای مدیریت صادرات پیشرفته محصولات"""
    
    def __init__(self, redis_url="redis://localhost:6379/0"):
        self.redis_manager = RedisExportManager(redis_url)
        self.chunk_size = 100
        self.LANGUAGE_FONTS = {
            'fa': 'B Nazanin',  # فارسی
            'ar': 'Arial',      # عربی
            'ru': 'Arial',      # روسی
            'zh': 'SimSun',     # چینی
            'en': 'Arial'       # انگلیسی
        }
    
    def export_products_to_excel(self, message, use_cache=True, background_processing=False):
        """
        صادرات کامل محصولات به اکسل
        """
        try:
            # بررسی subscription و پروفایل
            if not self._check_subscription(message):
                return {'error': 'اشتراک مورد نیاز است'}
            
            profile = self._get_user_profile(message)
            if not profile:
                return {'error': 'پروفایل یافت نشد'}
                
            if not profile.seller_mode:
                return {'error': 'حالت فروشنده مورد نیاز است'}
            
            store = self._get_user_store(profile)
            if not store:
                return {'error': 'فروشگاهی برای این پروفایل یافت نشد'}
            
            # دریافت زبان کاربر
            user_lang = profile.lang
            
            # بررسی کش
            if use_cache:
                cached_result = self.redis_manager.get_cached_export(store.id)
                if cached_result and cached_result.get('file_data'):
                    try:
                        file_data = cached_result['file_data']
                        if isinstance(file_data, str):
                            file_data = file_data.encode('utf-8')
                        
                        if file_data[:4] == b'PK\x03\x04':
                            file_buffer = BytesIO(file_data)
                            file_buffer.name = cached_result['filename']
                            
                            print(f"Using cached file - Size: {len(file_data)} bytes")
                            
                            return {
                                'file_buffer': file_buffer,
                                'filename': cached_result['filename'],
                                'from_cache': True,
                                'metadata': cached_result['metadata'],
                                'store_name': store.name,
                                'user_lang': user_lang
                            }
                        else:
                            print("Cached file has invalid Excel signature")
                    except Exception as e:
                        print(f"Error processing cached file: {e}")
            
            # تولید فایل جدید
            return self._generate_excel_file(store, user_lang)
            
        except Exception as e:
            print(f"Export error: {traceback.format_exc()}")
            return {'error': f'خطا در تولید فایل اکسل: {str(e)}'}
    
    def _check_subscription(self, message):
        """بررسی subscription"""
        try:
            return subscription.subscription_offer(message)
        except:
            return True
    
    def _get_user_profile(self, message):
        """دریافت پروفایل کاربر"""
        try:
            return ProfileModel.objects.get(tel_id=message.from_user.id)
        except ProfileModel.DoesNotExist:
            return None
    
    def _get_user_store(self, profile):
        """دریافت فروشگاه کاربر"""
        try:
            return Store.objects.get(owner=profile)
        except Store.DoesNotExist:
            return None
    
    def _get_store_statistics(self, store):
        """آمار فروشگاه"""
        stats = {
            'total_products': Product.objects.filter(store=store).count(),
            'active_products': Product.objects.filter(store=store, status=True).count(),
            'products_with_variants': Product.objects.filter(
                store=store, 
                variants__isnull=False
            ).distinct().count(),
            'total_variants': ProductVariant.objects.filter(product__store=store).count(),
            'categories_count': Category.objects.filter(store=store).count(),
            'products_with_discount': Product.objects.filter(store=store, discount__gt=0).count(),
            'products_with_images': Product.objects.filter(
                store=store
            ).exclude(main_image='').count(),
        }
        
        from django.db.models import F, Sum
        stock_value = Product.objects.filter(store=store).aggregate(
            total_value=Sum(F('price') * F('stock'))
        )
        stats['total_stock_value'] = float(stock_value['total_value'] or 0)
        
        return stats
    
    def _get_translation(self, key, lang):
        """دریافت ترجمه برای کلید و زبان مشخص"""
        return translations.get(key, {}).get(lang, translations.get(key, {}).get('en', key))
    
    def _get_font_for_language(self, lang):
        """دریافت فونت مناسب برای زبان"""
        return self.LANGUAGE_FONTS.get(lang, 'Arial')
    
    def _auto_adjust_column_widths(self, worksheet):
        """تنظیم خودکار عرض ستون‌ها بر اساس محتوای سلول‌ها"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        adjusted_length = cell_length * 1.2
                        if adjusted_length > max_length:
                            max_length = adjusted_length
                except:
                    pass
            
            adjusted_width = min(max(max_length, 8), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_center_alignment_to_sheet(self, worksheet):
        """اعمال وسط‌چین به تمام سلول‌های یک شیت"""
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
    
    def _generate_excel_file(self, store, user_lang):
        """تولید فایل اکسل با تنظیمات پیشرفته"""
        try:
            output = BytesIO()
            
            # ایجاد workbook جدید
            workbook = openpyxl.Workbook()
            
            # حذف sheet پیش‌فرض
            workbook.remove(workbook.active)
            
            # ایجاد sheets با نام‌های ترجمه شده
            products_sheet = workbook.create_sheet(self._get_translation('products_sheet', user_lang))
            variants_sheet = workbook.create_sheet(self._get_translation('variants_sheet', user_lang))
            summary_sheet = workbook.create_sheet(self._get_translation('summary_sheet', user_lang))
            categories_sheet = workbook.create_sheet(self._get_translation('categories_sheet', user_lang))
            
            # تنظیم فونت برای هر زبان
            font_name = self._get_font_for_language(user_lang)
            base_font = Font(name=font_name, size=10)
            header_font = Font(name=font_name, size=11, bold=True)
            
            # پر کردن داده‌ها با ترجمه و فونت مناسب
            self._fill_products_sheet(products_sheet, store, user_lang, base_font, header_font)
            self._fill_variants_sheet(variants_sheet, store, user_lang, base_font, header_font)
            self._fill_summary_sheet(summary_sheet, store, user_lang, base_font, header_font)
            self._fill_categories_sheet(categories_sheet, store, user_lang, base_font, header_font)
            
            # تنظیم خودکار عرض ستون‌ها برای همه sheets
            for sheet in workbook.worksheets:
                self._auto_adjust_column_widths(sheet)
                # اعمال وسط‌چین به تمام سلول‌های شیت
                self._apply_center_alignment_to_sheet(sheet)
            
            # ذخیره workbook در BytesIO
            workbook.save(output)
            output.seek(0)
            file_data = output.getvalue()
            
            # بررسی signature فایل
            if file_data[:4] != b'PK\x03\x04':
                raise ValueError("Generated file is not a valid Excel file")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{store.name}_products_export_{timestamp}.xlsx"
            
            # ذخیره در کش
            cached_data = {
                'file_data': file_data,
                'filename': filename,
                'metadata': self._get_store_statistics(store),
                'store_name': store.name,
                'generated_at': datetime.now().isoformat(),
                'user_lang': user_lang
            }
            
            try:
                self.redis_manager.cache_export(store.id, cached_data)
            except Exception as e:
                print(f"Cache error (non-critical): {e}")
                # خطای کش بحرانی نیست، ادامه بده
            
            # ایجاد BytesIO برای ارسال
            file_buffer = BytesIO(file_data)
            file_buffer.name = filename
            
            return {
                'file_buffer': file_buffer,
                'filename': filename,
                'from_cache': False,
                'metadata': cached_data['metadata'],
                'store_name': store.name,
                'user_lang': user_lang
            }
            
        except Exception as e:
            print(f"Excel generation error: {traceback.format_exc()}")
            raise e
    
    def _fill_products_sheet(self, sheet, store, user_lang, base_font, header_font):
        """پر کردن شیت محصولات"""
        headers = [
            self._get_translation('id', user_lang),
            self._get_translation('product_name', user_lang),
            self._get_translation('product_code', user_lang),
            self._get_translation('brand', user_lang),
            self._get_translation('category', user_lang),
            self._get_translation('base_price', user_lang),
            self._get_translation('discount_percent', user_lang),
            self._get_translation('final_price', user_lang),
            self._get_translation('stock', user_lang),
            self._get_translation('status', user_lang),
            self._get_translation('unit', user_lang),
            self._get_translation('min_quantity', user_lang),
            self._get_translation('max_quantity', user_lang),
            self._get_translation('quantity_step', user_lang),
            self._get_translation('images_count', user_lang),
            self._get_translation('variants_count', user_lang),
            self._get_translation('description', user_lang),
            self._get_translation('has_main_image', user_lang)
        ]
        
        # اضافه کردن headers با فونت header و وسط‌چین
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
        
        # اضافه کردن داده‌ها با وسط‌چین
        row_num = 2
        for product_data in self._stream_products_data(store):
            # ترجمه مقادیر وضعیت
            status_text = self._get_translation('active', user_lang) if product_data['status'] == 'فعال' else self._get_translation('inactive', user_lang)
            has_image_text = self._get_translation('yes', user_lang) if product_data['has_main_image'] else self._get_translation('no', user_lang)
            
            data_row = [
                product_data['id'],
                product_data['name'],
                product_data['code'],
                product_data['brand'],
                product_data['category_full_path'],
                product_data['price'],
                product_data['discount'],
                product_data['final_price'],
                product_data['stock'],
                status_text,
                f"{product_data['unit_name']} ({product_data['unit_symbol']})",
                product_data['min_quantity'],
                product_data['max_quantity'] or '',
                product_data['quantity_step'],
                product_data['images_count'],
                len(product_data['variants']),
                product_data['description'],
                has_image_text
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = sheet.cell(row=row_num, column=col, value=value)
                cell.font = base_font
                cell.alignment = center_alignment
            
            row_num += 1
    
    def _fill_variants_sheet(self, sheet, store, user_lang, base_font, header_font):
        """پر کردن شیت واریانت‌ها"""
        headers = [
            self._get_translation('product_id', user_lang),
            self._get_translation('product_name', user_lang),
            self._get_translation('variant_id', user_lang),
            self._get_translation('sku', user_lang),
            self._get_translation('attributes', user_lang),
            self._get_translation('price_override', user_lang),
            self._get_translation('final_price', user_lang),
            self._get_translation('stock', user_lang),
            self._get_translation('status', user_lang)
        ]
        
        # اضافه کردن headers با فونت header و وسط‌چین
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
        
        # اضافه کردن داده‌ها با وسط‌چین
        row_num = 2
        for product_data in self._stream_products_data(store):
            for variant in product_data['variants']:
                # ترجمه وضعیت
                status_text = self._get_translation('active', user_lang) if variant['is_active'] else self._get_translation('inactive', user_lang)
                
                data_row = [
                    product_data['id'],
                    product_data['name'],
                    variant['variant_id'],
                    variant['sku'],
                    variant['values'],
                    variant['price_override'] or '',
                    variant['final_price'],
                    variant['stock'],
                    status_text
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = sheet.cell(row=row_num, column=col, value=value)
                    cell.font = base_font
                    cell.alignment = center_alignment
                
                row_num += 1
    
    def _fill_summary_sheet(self, sheet, store, user_lang, base_font, header_font):
        """پر کردن شیت خلاصه"""
        stats = self._get_store_statistics(store)
        
        summary_data = [
            [self._get_translation('store_name', user_lang), store.name],
            [self._get_translation('store_address', user_lang), f"{store.address}, {store.city}"],
            [self._get_translation('export_date', user_lang), datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            [self._get_translation('total_products', user_lang), stats['total_products']],
            [self._get_translation('active_products', user_lang), stats['active_products']],
            [self._get_translation('inactive_products', user_lang), stats['total_products'] - stats['active_products']],
            [self._get_translation('products_with_variants', user_lang), stats['products_with_variants']],
            [self._get_translation('total_variants', user_lang), stats['total_variants']],
            [self._get_translation('categories_count', user_lang), stats['categories_count']],
            [self._get_translation('products_with_discount', user_lang), stats['products_with_discount']],
            [self._get_translation('products_with_images', user_lang), stats['products_with_images']],
            [self._get_translation('total_stock_value', user_lang), f"{stats['total_stock_value']:,.0f}"],
        ]
        
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row, (metric, value) in enumerate(summary_data, 1):
            # سلول متریک
            metric_cell = sheet.cell(row=row, column=1, value=metric)
            metric_cell.font = header_font
            metric_cell.alignment = center_alignment
            
            # سلول مقدار
            value_cell = sheet.cell(row=row, column=2, value=value)
            value_cell.font = base_font
            value_cell.alignment = center_alignment
    
    def _fill_categories_sheet(self, sheet, store, user_lang, base_font, header_font):
        """پر کردن شیت دسته‌بندی‌ها"""
        headers = [
            self._get_translation('category_id', user_lang),
            self._get_translation('category_name', user_lang),
            self._get_translation('full_path', user_lang),
            self._get_translation('total_products', user_lang),
            self._get_translation('active_products', user_lang),
            self._get_translation('parent_category', user_lang),
            self._get_translation('position', user_lang),
            self._get_translation('status', user_lang)
        ]
        
        # اضافه کردن headers با فونت header و وسط‌چین
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
        
        # اضافه کردن داده‌ها با وسط‌چین
        row_num = 2
        for category in Category.objects.filter(store=store):
            product_count = Product.objects.filter(category=category, store=store).count()
            active_count = Product.objects.filter(category=category, store=store, status=True).count()
            
            # ترجمه وضعیت
            status_text = self._get_translation('active', user_lang) if category.status else self._get_translation('inactive', user_lang)
            
            data_row = [
                category.id,
                category.title,
                category.get_full_path(),
                product_count,
                active_count,
                category.parent.title if category.parent else self._get_translation('no', user_lang),
                category.position,
                status_text
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = sheet.cell(row=row_num, column=col, value=value)
                cell.font = base_font
                cell.alignment = center_alignment
            
            row_num += 1
    
    def _stream_products_data(self, store):
        """جریان داده محصولات"""
        products_query = Product.objects.filter(store=store).select_related(
            'category', 'unit'
        ).prefetch_related('images').order_by('category__title', 'name')
        
        total_count = products_query.count()
        
        for offset in range(0, total_count, self.chunk_size):
            products_chunk = products_query[offset:offset + self.chunk_size]
            
            product_ids = [p.id for p in products_chunk]
            variants_map = {}
            
            variants = ProductVariant.objects.filter(
                product_id__in=product_ids
            ).prefetch_related('values__option')
            
            for variant in variants:
                if variant.product_id not in variants_map:
                    variants_map[variant.product_id] = []
                variants_map[variant.product_id].append(variant)
            
            for product in products_chunk:
                product_data = self._serialize_product(product)
                
                if product.id in variants_map:
                    variants_data = [
                        self._serialize_variant(variant) 
                        for variant in variants_map[product.id]
                    ]
                    product_data['variants'] = variants_data
                else:
                    product_data['variants'] = []
                
                yield product_data
            
            del products_chunk
            del variants_map
    
    def _serialize_product(self, product):
        """سریالایز کردن داده محصول"""
        return {
            'id': product.id,
            'name': product.name,
            'code': product.code,
            'brand': product.brand or '',
            'category_full_path': product.category.get_full_path() if product.category else '',
            'category_name': product.category.title if product.category else '',
            'price': float(product.price),
            'discount': float(product.discount),
            'final_price': float(product.final_price),
            'stock': product.stock,
            'status': 'فعال' if product.status else 'غیرفعال',
            'unit_name': product.unit.name if product.unit else '',
            'unit_symbol': product.unit.symbol if product.unit else '',
            'min_quantity': float(product.min_quantity),
            'max_quantity': float(product.max_quantity) if product.max_quantity else None,
            'quantity_step': float(product.quantity_step),
            'description': self._truncate_description(product.description),
            'has_main_image': bool(product.main_image),
            'images_count': product.images.count(),
        }
    
    def _serialize_variant(self, variant):
        """سریالایز کردن داده واریانت"""
        variant_values = " | ".join([
            f"{value.option.name}:{value.value}" 
            for value in variant.values.all()
        ])
        
        return {
            'variant_id': variant.id,
            'sku': variant.sku or '',
            'values': variant_values,
            'price_override': float(variant.price_override) if variant.price_override else None,
            'final_price': float(variant.final_price),
            'stock': variant.stock,
            'is_active': variant.product.status if variant.product else True
        }
    
    def _truncate_description(self, description, max_length=100):
        """کوتاه کردن توضیحات"""
        if not description:
            return ''
        return (description[:max_length] + '...') if len(description) > max_length else description
    
    def clear_cache_for_store(self, store_id):
        """پاک کردن کش برای یک فروشگاه خاص"""
        cache_key = self.redis_manager._make_export_key(store_id)
        self.redis_manager.redis_client.delete(cache_key)



################## VIDEO PROMPTS ################

class UltraVideoPrompter:
    CACHE_CHANNEL = "@Botshop_trainer"

    def __init__(self, command: str, skip_data: str = "skip_video", app=None):
        self.command = command
        self.skip_data = f"{skip_data}|{command}"
        self.app = app

    # ----------------------------------------------------------------
    def __call__(self, func):

        @functools.wraps(func)
        def wrapper(message, *args, **kwargs):

            print("\n==============================")
            print("=== UltraVideoPrompter wrapper executed ===")


            bot = self.app or app
            user_id = message.chat.id

            print("User ID =", user_id)

            # ---------------------------------------------------------
            # 1) گرفتن پروفایل
            try:
                profile = ProfileModel.objects.get(tel_id=user_id)
                lang = profile.lang
                print("Profile Loaded:", profile)
            except ProfileModel.DoesNotExist:
                profile = None
                lang = "en"
                print("Profile NOT FOUND -> using 'en'")

            # ---------------------------------------------------------
            # 2) چاپ وضعیت پروفایل
            print("Language:", lang)
            print("hidden_videos =", profile.hidden_videos if profile else None)
            print("command =", self.command)

            # ---------------------------------------------------------
            # 3) اگر hide فعال است → skip کن
            if profile:
                hidden_flag = profile.hidden_videos.get(self.command)
                print("HIDDEN FLAG =", hidden_flag)

                if hidden_flag:
                    print(">>> SKIP VIDEO (flag=True)")
                    print("==============================\n")
                    return func(message, *args, **kwargs)

            # ---------------------------------------------------------
            # 4) مسیر ویدیو
            video_path = Path(
                settings.BASE_DIR,
                f"media/promptvideos/{lang}/{self.command}.mp4"
            )

            print("Video path:", video_path)
            print("Exists:", os.path.exists(video_path))

            # ---------------------------------------------------------
            # 5) ارسال
            self._smart_send(message, bot, video_path, lang)

            print("==============================\n")
            return func(message, *args, **kwargs)

        return wrapper

    # ----------------------------------------------------------------
    def _smart_send(self, message, bot, video_path, lang):
        user_id = message.chat.id
        print(">>> Enter _smart_send()")

        markup = self._build_button()

        # ---------------------------------------------------------
        # 1) بررسی cache
        cached = CachedMedia.objects.filter(video_path=str(video_path)).first()
        print("Cached record:", cached)

        if cached:
            print("Cached file_id:", cached.file_id)

        # ---------------------------------------------------------
        # اگر cache موجود باشد
        if cached and cached.file_id:
            print(">>> Sending from CACHE to user")
            try:
                bot.send_video(
                    user_id,
                    cached.file_id,
                    caption="🎬 راهنما",
                    reply_markup=markup,
                    parse_mode="HTML"
                )
            except Exception as e:
                print("ERROR sending cached file:", e)
            return

        # ---------------------------------------------------------
        # 2) فایل وجود ندارد؟
        if not os.path.exists(video_path):
            print("❌ ERROR: Video file does NOT exist!")
            bot.send_message(user_id, "❌ ویدیو یافت نشد.")
            return

        # ---------------------------------------------------------
        # 3) ارسال به کانال
        print(">>> Uploading NEW video to CACHE CHANNEL...")

        try:
            with open(video_path, "rb") as f:
                sent = bot.send_video(
                    chat_id=self.CACHE_CHANNEL,
                    video=f,
                    caption=f"{lang}-{self.command}.mp4"
                )
            print("Upload success!")
        except Exception as e:
            print("❌ ERROR uploading to channel:", e)
            bot.send_message(user_id, "❌ خطا در ارسال ویدیو به کانال اصلی.")
            return

        # ---------------------------------------------------------
        # 4) ذخیره cache
        try:
            cache_record = CachedMedia.objects.create(
                video_path=str(video_path),
                file_id=sent.video.file_id,
                channel_message_id=sent.message_id,
                media_type="video"
            )
            print("CachedMedia SAVED:", cache_record)
        except Exception as e:
            print("❌ ERROR saving CachedMedia:", e)

        # ---------------------------------------------------------
        # 5) ارسال نهایی به کاربر
        print(">>> Sending NEW video to user")
        try:
            bot.send_video(
                user_id,
                sent.video.file_id,
                caption="🎬 راهنما",
                reply_markup=markup,
                parse_mode="HTML"
            )
        except Exception as e:
            print("❌ ERROR sending final video to user:", e)

        print(">>> _smart_send() finished")

    # ----------------------------------------------------------------
    def _build_button(self):
        markup = types.InlineKeyboardMarkup()
        btn = types.InlineKeyboardButton(
            "🔕 دیگر نمایش داده نشود", callback_data=self.skip_data
        )
        markup.add(btn)
        return markup

    # ----------------------------------------------------------------
    @staticmethod
    def handle_skip_callback(call):
        bot = app
        user_id = call.message.chat.id

        print(">>> handle_skip_callback()")
        print("callback data:", call.data)

        try:
            profile = ProfileModel.objects.get(tel_id=user_id)
        except ProfileModel.DoesNotExist:
            print("Profile not found")
            return bot.answer_callback_query(call.id, "❌ پروفایل یافت نشد")

        if "|" not in call.data:
            print("Invalid callback data")
            return bot.answer_callback_query(call.id, "❌ خطا")

        _, command = call.data.split("|", 1)
        print("Skip command =", command)

        profile.hidden_videos[command] = True
        profile.save(update_fields=['hidden_videos'])

        print("Hidden updated:", profile.hidden_videos)

        bot.answer_callback_query(
            call.id,
            t(call.message, "video_hidden")
        )



class SendPhotoWithMarkup(SendMarkup):
    def __init__(self,bot,chat_id,photo_path=None,photo_url=None,file_id=None,caption=None,buttons=None,button_layout=None,handlers=None):
        super().__init__(bot, chat_id, caption, buttons, button_layout, handlers)
        self.photo_path = photo_path
        self.photo_url = photo_url
        self.file_id = file_id
        
    def send(self):
        """ارسال عکس با caption و keyboard"""
        try:
            markup = self.generate_keyboard()
            
            # ارسال عکس از مسیر فایل
            if self.file_id:
                self.bot.send_photo(
                    chat_id=self.chat_id,
                    photo=self.file_id,
                    caption=self.text,
                    reply_markup=markup,
                    parse_mode="HTML"
                )

            elif self.photo_path:
                with open(self.photo_path, 'rb') as photo:
                    self.bot.send_photo(
                        chat_id=self.chat_id,
                        photo=photo,
                        caption=self.text,
                        reply_markup=markup,
                        parse_mode="HTML"
                    )

            elif self.photo_url:
                self.bot.send_photo(
                    chat_id=self.chat_id,
                    photo=self.photo_url,
                    caption=self.text,
                    reply_markup=markup,
                    parse_mode="HTML"
                )

            else:
                print("هیچ منبع عکسی مشخص نشده است")
                
        except Exception as e:
            print(f"Error in SendPhotoWithMarkup.send: {traceback.format_exc()}")
            # تلاش برای ارسال بدون عکس در صورت خطا
            try:
                super().send()
            except Exception as e2:
                print(f"Error sending without photo: {e2}")
    
    def edit(self, message_id):
        """ویرایش عکس و caption"""
        try:
            markup = self.generate_keyboard()
            
            # ویرایش caption عکس
            self.bot.edit_message_caption(
                chat_id=self.chat_id,
                message_id=message_id,
                caption=self.text,
                reply_markup=markup,
                parse_mode="HTML"
            )
        except Exception as e:
            if "message is not modified" not in str(e):
                print(f"Error in SendPhotoWithMarkup.edit: {traceback.format_exc()}")
                # تلاش برای ویرایش بدون keyboard
                try:
                    self.bot.edit_message_caption(
                        chat_id=self.chat_id,
                        message_id=message_id,
                        caption=self.text,
                        parse_mode="HTML"
                    )
                except Exception as e2:
                    print(f"Error editing without keyboard: {e2}")


def t(msg, key, chat_id=None, profile=None, lang=None, **kwargs):
    try:
        if isinstance(msg, types.Message):
            message = msg
        elif isinstance(msg, types.CallbackQuery):
            message = msg.message
        else:
            message = None

        if chat_id is None and message:
            chat_id = message.chat.id

        # استفاده از profile یا lang اگر داده شده
        if profile:
            lang = profile.lang
        elif lang is None and chat_id:
            lang = ProfileModel.objects.get(tel_id=chat_id).lang

        text = translations.get(key, {}).get(lang, translations.get(key, {}).get("en", key))

        if kwargs:
            text = text.format(**kwargs)

        return text
    except Exception:
        print(traceback.format_exc())
        return key


# @add_performance_monitoring_to_class
class SendStore:
    """
    Optimized & Stateless Store Sender
    - Only ONE DB query for Profile + Store
    - No duplicated logic
    - Clean separation of concerns
    """

    def __init__(self, bot: TeleBot):
        self.bot = bot

    # =============================
    # Public API
    # =============================
    def show_store_info(self, message: Message):
        try:
            chat_id = message.chat.id

            profile, store = self._load_context(chat_id)

            buttons = self._generate_buttons(profile, store)

            logo = self._get_logo_source(profile, store)

            SendPhotoWithMarkup(
                bot=self.bot,
                chat_id=chat_id,
                photo_path=logo["value"] if logo["type"] == "path" else None,
                file_id=logo["value"] if logo["type"] == "file_id" else None,
                caption=self._generate_caption(profile, store),
                buttons=buttons,
                button_layout=[2, 2, 2, 2, 2, 1],
                handlers=self._generate_handlers(profile)
            ).send()


        except Exception:
            print(traceback.format_exc())

    # =============================
    # Context Loader (ONE DB HIT)
    # =============================
    def _load_context(self, chat_id):
        profile = ProfileModel.objects.get(tel_id=chat_id)

        store = (
            Store.objects
            .select_related("owner")
            .prefetch_related("store_address")
            .filter(owner=profile)
            .first()
        )

        return profile, store

    # =============================
    # UI Builders (NO DB)
    # =============================
    def _generate_buttons(self, profile, store):
        buttons = {}

        if not store:
            session = session_manager.get_user_session(profile.tel_id, namespace="createshop")
            text = f"{t('message','store_name', profile=profile)}: {session.get("take_name_d") if session.get("take_name_d" or None) else ' --- '}"
            buttons[f"{text}"] = {
                "callback_data": "store_name", "index": 1
            }

            buttons[t(
                'message', 'address',
                profile=profile,
                address_text=' --- '
            )] = {"callback_data": "noop", "index": 1}

            buttons[t(
                'message', 'store_description',
                profile=profile,
            )] = {"callback_data": "store_description", "index": 1}

            buttons[t(
                'message', 'store_payment_metod',
                profile=profile,
            )] = {"callback_data": "store_payment_metod", "index": 1}

            buttons[t(
                'message', 'store_telegram_channel',
                profile=profile,
            )] = {"callback_data": "store_telegram_channel", "index": 1}


            text = t('message', 'set_store_logo', profile=profile) if not session.get("take_logo_d") else t('message', 'change_store_logo', profile=profile)

            buttons[f"{text}"] = {"callback_data": "set_store_logo", "index": 1}

            buttons[t('message', 'submit_information', profile=profile)] = {
                "callback_data": "submit_info", "index": 1
            }

            return buttons

        addr = store.get_address()

        addr_text = self._format_address(addr, profile.lang)

        buttons[f"{t('message','store_name', profile=profile)}: {store.name}"] = {
            "callback_data": "store_name", "index": 1
        }

        buttons[t('message', 'address', profile=profile, address_text=addr_text)] = {
            "callback_data": "buy_product", "index": 1
        }

        buttons[t(
            'message', 'store_description',
            profile=profile,
        )] = {"callback_data": "store_description", "index": 1}

        buttons[t(
            'message', 'store_payment_metod',
            profile=profile,
        )] = {"callback_data": "store_payment_metod", "index": 1}

        buttons[t(
            'message', 'store_telegram_channel',
            profile=profile,
        )] = {"callback_data": "store_telegram_channel", "index": 1}


        buttons[f"{t('message', 'change_store_logo', profile=profile)}"] = {"callback_data": "set_store_logo", "index": 1}

        buttons[t('message', 'store_delete', profile=profile)] = {
                "callback_data": "submit_info", "index": 1
            }

        return buttons

    def _generate_caption(self, profile, store):
        session = session_manager.get_user_session(profile.tel_id, namespace="createshop")
        if store:
            caption = f"<b>{store.name}</b> \n\n{store.description}"
        elif session.get("take_description_d" or None):
            if session.get("take_name_d"):
                caption = f"{session.get('take_name_d')} \n\n"
            caption += session.get("take_description_d" or None)
        else: 
            caption = t("message", "store_setup_info", profile=profile)
 
        return caption

    def _generate_handlers(self, profile):
        if profile.lang == "fa":
            msg = "خرید با موفقیت ثبت شد!"
        else:
            msg = "Purchase completed successfully!"

        return {
            "buy_product": lambda call: self.bot.answer_callback_query(call.id, msg)
        }

    # =============================
    # Helpers (Pure Logic)
    # =============================
    def _get_logo_source(self, profile, store):
        session = session_manager.get_user_session(profile.tel_id, namespace="createshop")

        if store and store.logo:
            return {"type": "path", "value": store.logo.path}

        if session.get("take_logo_d"):
            return {"type": "file_id", "value": session["take_logo_d"]}

        from django.conf import settings
        default_logo_path = os.path.join(
            settings.MEDIA_ROOT,
            "store_logos",
            f"{profile.lang}-default-store-logo.png"
        )

        return {"type": "path", "value": default_logo_path}



    @staticmethod
    def _format_address(addr, lang):
        if not addr:
            return "---"

        if lang == "fa":
            return f"{addr.shipping_country_name}، {addr.shipping_province_name}، {addr.shipping_city_name} ..."
        else:
            return f"{addr.shipping_city_name}, {addr.shipping_province_name}, {addr.shipping_country_name} ..."


    def take_name(self, call):
        try:
            session = session_manager.get_user_session(call.message.chat.id, namespace="createshop")
            profile, store = self._load_context(call.message.chat.id)
            text = t('message','enter_store_name', profile=profile)
            self.bot.delete_message(call.message.chat.id, call.message.message_id)
            try:
                self.bot.delete_message(call.message.chat.id, session.get("msg_id"))
            except:
                pass
            cancel_text = t("message", "cancel_action", profile=profile)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(types.KeyboardButton(cancel_text))
            self.bot.send_message(call.message.chat.id, text, reply_markup=markup)
            session["take_name"] = True
            session["take_data"] = True
            session_manager.set_user_session(call.message.chat.id, session, namespace="createshop")

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in take_name: {e}\n{error_details}")


    def take_logo(self, call):
        try:
            session = session_manager.get_user_session(call.message.chat.id, namespace="createshop")
            profile, store = self._load_context(call.message.chat.id)
            text = t('message','get_store_logo', profile=profile)
            self.bot.delete_message(call.message.chat.id, call.message.message_id)
            try:
                self.bot.delete_message(call.message.chat.id, session.get("msg_id"))
            except:
                pass
            cancel_text = t("message", "cancel_action", profile=profile)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(types.KeyboardButton(cancel_text))
            self.bot.send_message(call.message.chat.id, text, reply_markup=markup)
            session["take_logo"] = True
            session["take_data"] = True
            session_manager.set_user_session(call.message.chat.id, session, namespace="createshop")

        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Error in take_name: {e}\n{error_details}")

    def take_description(self, call):
        try:
            session = session_manager.get_user_session(call.message.chat.id, namespace="createshop")
            profile, store = self._load_context(call.message.chat.id)
            text = t('message','writ_store_description', profile=profile)
            self.bot.delete_message(call.message.chat.id, call.message.message_id)
            try:
                self.bot.delete_message(call.message.chat.id, session.get("msg_id"))
            except:
                pass
            cancel_text = t("message", "cancel_action", profile=profile)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(types.KeyboardButton(cancel_text))
            self.bot.send_message(call.message.chat.id, text, reply_markup=markup)
            session["take_description"] = True
            session["take_data"] = True
            session_manager.set_user_session(call.message.chat.id, session, namespace="createshop")

        except:
            print(traceback.format_exc())


    def take_telegram_channel(self, call):
        try:
            session = session_manager.get_user_session(call.message.chat.id, namespace="createshop")
            profile, store = self._load_context(call.message.chat.id)
            text = t('message','enter_channel_id', profile=profile)
            self.bot.delete_message(call.message.chat.id, call.message.message_id)
            try:
                self.bot.delete_message(call.message.chat.id, session.get("msg_id"))
            except:
                pass
            cancel_text = t("message", "cancel_action", profile=profile)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(types.KeyboardButton(cancel_text))
            self.bot.send_message(call.message.chat.id, text, reply_markup=markup)
            session["take_telegram_channel"] = True
            session["take_data"] = True
            session_manager.set_user_session(call.message.chat.id, session, namespace="createshop")

        except:
            print(traceback.format_exc())

